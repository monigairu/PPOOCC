"""
Excel シート構造スキャナー

シート内のラベルセルを検出し、対応する入力候補セルとの
マッピングを自動生成する。
"""
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.workbook.workbook import Workbook


def scan_label_cells(
    workbook: Workbook,
    sheet_name: str,
) -> dict[str, list[str]]:
    """
    シートを全スキャンして、ラベル名と候補入力セルの対応マップを生成する。

    ラベルセル（文字列）を見つけたら、その右隣・下隣のセルを
    候補入力セルとして登録する。同じラベルが複数箇所にある場合も
    すべて検出する。

    Args:
        workbook: 対象の Workbook オブジェクト
        sheet_name: スキャン対象のシート名

    Returns:
        {ラベル名: [候補セルアドレスのリスト]}
        例: {"炉型": ["C7", "G9", "K9"]}
    """
    sheet = workbook[sheet_name]
    label_map: dict[str, list[str]] = {}

    for row in sheet.iter_rows():
        for cell in row:
            # 結合セルはスキップ（左上セル以外は読めない）
            if isinstance(cell, MergedCell):
                continue

            # 文字列セルのみラベル候補とする
            if not isinstance(cell.value, str):
                continue

            label = cell.value.strip()
            if not label:
                continue

            candidates = []

            # 右隣のセルをチェック
            right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
            if _is_writable_input_cell(right_cell):
                candidates.append(right_cell.coordinate)

            # 下隣のセルをチェック
            below_cell = sheet.cell(row=cell.row + 1, column=cell.column)
            if _is_writable_input_cell(below_cell):
                candidates.append(below_cell.coordinate)

            # 候補が見つかった場合のみマップに追加
            if candidates:
                if label not in label_map:
                    label_map[label] = []
                for candidate in candidates:
                    if candidate not in label_map[label]:
                        label_map[label].append(candidate)

    return label_map


def _is_writable_input_cell(cell) -> bool:
    """
    書き込み可能な入力セルの候補かどうかを判定する。

    結合セルは書き込めないので除外する。
    空セル または 数値が入っているセルを入力候補とみなす。

    Args:
        cell: openpyxl の Cell オブジェクト

    Returns:
        書き込み可能な入力セル候補なら True
    """
    # 結合セルは書き込み不可なので除外
    if isinstance(cell, MergedCell):
        return False

    value = cell.value
    if value is None:
        return True
    if isinstance(value, (int, float)):
        return True
    return False