"""
Excel ファイル入出力モジュール

このモジュールは Excel ファイルの読込・保存・コピーといった
基本的な入出力操作を提供する。
"""
import shutil

from openpyxl import load_workbook, Workbook


def copy_excel_file(src: str, dst: str) -> None:
    """
    Excel ファイルをコピーする。

    テンプレートを直接編集せず、コピーした作業用ファイルに対して
    書き込みを行うために使用する。

    Args:
        src: コピー元ファイルのパス
        dst: コピー先ファイルのパス
    """
    shutil.copy2(src, dst)


def load_workbook_file(path: str) -> Workbook:
    """
    Excel ファイルを読み込んで Workbook オブジェクトを返す。

    Args:
        path: 読み込む Excel ファイルのパス

    Returns:
        openpyxl の Workbook オブジェクト
    """
    return load_workbook(path)


def save_workbook_file(workbook: Workbook, path: str) -> None:
    """
    Workbook をファイルとして保存する。

    Args:
        workbook: 保存する Workbook オブジェクト
        path: 保存先ファイルのパス
    """
    workbook.save(path)