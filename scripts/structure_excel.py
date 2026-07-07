"""
汎用 Excel 構造化スクリプト

入力資料の様式に依存せず、任意の Excel シートから「1レコード=1行（複数行に
またがる場合は結合）」の構造化データ (JSON / CSV) を生成する。

対応している一般的なレイアウト:
- セル結合（結合範囲の左上値を全セルへ展開してから解析）
- 複数行ヘッダー（縦結合・横結合からヘッダー行数を自動推定し「親/子」で連結）
- 1項目が複数行にまたがる明細（アンカー列に値が現れた行をレコード開始とみなす）
- 同一シート内の複数テーブル（空白で分離された領域をブロックとして自動分割）
- ガントチャート等の塗りつぶし表現（--fills で塗りセルをヘッダー名にマップ）

使い方:
    uv run python scripts/structure_excel.py <xlsx> [options]

    --sheet NAME    対象シート名（省略時は全シート）
    --range A1:N99  解析範囲を限定（隣接する別表を切り離したい場合）
    --fills         塗りつぶしセルを marks として抽出（工程表向け）
    --list          検出したブロック一覧のみ表示して終了
    -o out.json     JSON 出力先（省略時は標準出力）
    --csv DIR       ブロックごとの CSV も DIR に出力

備考:
- 数式セルはキャッシュ値（最後に Excel が計算した値）を読む。
- 特定の様式・費目へのハードコードはしない（横断ルール準拠）。
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import deque
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

# ブロック検出時の許容ギャップ（この範囲内の非空セル同士を同一ブロックとみなす）
ROW_GAP = 3  # 明細様式は項目行と数値行の間に空行が入ることがあるため広め
COL_GAP = 1  # 空列1つで別テーブルと判断する

# 塗りつぶし抽出で「マークなし」とみなす色
UNMARKED_RGB = {None, "FFFFFFFF"}


def build_grid(ws: Worksheet) -> tuple[dict[tuple[int, int], object], dict[tuple[int, int], tuple[int, int]]]:
    """セル値のグリッドと、結合セル→左上座標のマップを構築する。

    Returns:
        (grid, merge_origin)
        grid: {(row, col): 値} 結合範囲は左上値を全セルへ展開済み
        merge_origin: {(row, col): (左上row, 左上col)} 結合範囲内のセルのみ
    """
    grid: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                grid[(cell.row, cell.column)] = cell.value

    merge_origin: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in ws.merged_cells.ranges:
        top_left = (rng.min_row, rng.min_col)
        value = grid.get(top_left)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merge_origin[(r, c)] = top_left
                if value is not None:
                    grid[(r, c)] = value
    return grid, merge_origin


def collect_presence(ws: Worksheet, grid: dict[tuple[int, int], object]) -> set[tuple[int, int]]:
    """ブロック検出用の「セルが存在する」座標集合（値あり or 罫線あり）を返す。

    日本語帳票は値が疎でも罫線で表領域が定義されていることが多いため、
    罫線セルも領域の一部として扱うことで表の分断を防ぐ。
    """
    presence = set(grid.keys())
    for row in ws.iter_rows():
        for cell in row:
            b = cell.border
            if b and (
                (b.left and b.left.style)
                or (b.right and b.right.style)
                or (b.top and b.top.style)
                or (b.bottom and b.bottom.style)
            ):
                presence.add((cell.row, cell.column))
    return presence


def detect_blocks(
    presence: set[tuple[int, int]], grid: dict[tuple[int, int], object]
) -> list[tuple[int, int, int, int]]:
    """セル存在集合の連結成分からテーブルブロック（バウンディングボックス）を検出する。

    許容ギャップ内のセルを連結し、さらにボックス同士が重なるものは統合する。
    値を1つも含まないブロック（罫線のみの空枠）は除外する。

    Returns:
        [(min_row, min_col, max_row, max_col), ...] 左上から順
    """
    cells = presence
    visited: set[tuple[int, int]] = set()
    boxes: list[list[int]] = []

    for start in cells:
        if start in visited:
            continue
        queue = deque([start])
        visited.add(start)
        box = [start[0], start[1], start[0], start[1]]
        while queue:
            r, c = queue.popleft()
            box[0] = min(box[0], r)
            box[1] = min(box[1], c)
            box[2] = max(box[2], r)
            box[3] = max(box[3], c)
            for dr in range(-ROW_GAP, ROW_GAP + 1):
                for dc in range(-COL_GAP, COL_GAP + 1):
                    nb = (r + dr, c + dc)
                    if nb in cells and nb not in visited:
                        visited.add(nb)
                        queue.append(nb)
        boxes.append(box)

    # バウンディングボックスが重なる成分を統合（疎な列の孤立を防ぐ）
    merged = True
    while merged:
        merged = False
        for i in range(len(boxes)):
            for j in range(i + 1, len(boxes)):
                a, b = boxes[i], boxes[j]
                if a[0] <= b[2] and b[0] <= a[2] and a[1] <= b[3] and b[1] <= a[3]:
                    boxes[i] = [min(a[0], b[0]), min(a[1], b[1]), max(a[2], b[2]), max(a[3], b[3])]
                    boxes.pop(j)
                    merged = True
                    break
            if merged:
                break

    boxes = [
        b for b in boxes
        if any(b[0] <= r <= b[2] and b[1] <= c <= b[3] for (r, c) in grid)
    ]
    return sorted((tuple(b) for b in boxes), key=lambda b: (b[0], b[1]))


def split_preamble(
    box: tuple[int, int, int, int],
    grid: dict[tuple[int, int], object],
    merge_origin: dict[tuple[int, int], tuple[int, int]],
    merges,
) -> tuple[int, list[str]]:
    """ブロック先頭のタイトル・前文行を分離し、ヘッダー開始行とメタ情報を返す。

    「値の種類が少ない行」または「ブロック幅の大半を占める横結合がある行」を
    前文とみなす（表タイトル・工事件名・工期などが該当）。
    """
    min_row, min_col, max_row, max_col = box
    width = max_col - min_col + 1
    meta: list[str] = []

    row = min_row
    while row <= max_row:
        origins: list[tuple[int, int]] = []
        for c in range(min_col, max_col + 1):
            key = merge_origin.get((row, c), (row, c))
            if key in grid and key not in origins:
                origins.append(key)
        values = [grid[k] for k in origins]
        widest = 0
        for rng in merges:
            if rng.min_row <= row <= rng.max_row:
                span = min(rng.max_col, max_col) - max(rng.min_col, min_col) + 1
                widest = max(widest, span)
        is_preamble = len(values) < min(3, width) or widest >= width * 0.4
        if values and not is_preamble:
            break
        if values:
            meta.append(" | ".join(str(v).replace("\n", " ") for v in values))
        row += 1

    if row > max_row:  # 全行が前文扱いになった場合は先頭行から解析にフォールバック
        return min_row, []
    return row, meta


def detect_header_depth(
    box: tuple[int, int, int, int],
    grid: dict[tuple[int, int], object],
    merge_origin: dict[tuple[int, int], tuple[int, int]],
    merges,
) -> int:
    """ブロック先頭のヘッダー行数を結合セルの形状から推定する。

    - ヘッダー行から下へ伸びる縦結合があれば、その分ヘッダーを拡張
    - ヘッダー行に横結合があり、直下の行が文字列主体なら子ヘッダー行とみなす
    """
    min_row, min_col, max_row, max_col = box
    depth = 1
    changed = True
    while changed and min_row + depth - 1 < max_row:
        changed = False
        header_rows = range(min_row, min_row + depth)
        for rng in merges:
            if rng.min_col > max_col or rng.max_col < min_col:
                continue
            if rng.min_row in header_rows and rng.max_row >= min_row + depth:
                depth = rng.max_row - min_row + 1
                changed = True

    # 横結合の直下に子ヘッダー行（親見出しを分割する文字列群）があれば1行だけ拡張
    header_rows = range(min_row, min_row + depth)
    horizontal_cols: set[int] = set()
    for rng in merges:
        if rng.min_row in header_rows and rng.max_col > rng.min_col:
            horizontal_cols.update(range(max(rng.min_col, min_col), min(rng.max_col, max_col) + 1))
    if horizontal_cols and min_row + depth <= max_row:
        next_row = min_row + depth
        vals = {c: grid.get((next_row, c)) for c in range(min_col, max_col + 1)}
        vals = {c: v for c, v in vals.items() if v is not None}
        under = [c for c in vals if c in horizontal_cols]
        if (
            len(vals) >= 2
            and all(isinstance(v, str) for v in vals.values())
            and len(under) >= len(vals) * 0.6
        ):
            depth += 1
    return depth


def composite_headers(
    box: tuple[int, int, int, int], depth: int, grid: dict[tuple[int, int], object]
) -> dict[int, str]:
    """複数ヘッダー行を「親/子」形式で列ごとに連結する。ヘッダーが無い列は列記号。"""
    min_row, min_col, _, max_col = box
    headers: dict[int, str] = {}
    for c in range(min_col, max_col + 1):
        parts: list[str] = []
        for r in range(min_row, min_row + depth):
            v = grid.get((r, c))
            if v is None:
                continue
            s = str(v).replace("\n", " ").strip()
            if s and (not parts or parts[-1] != s):
                parts.append(s)
        headers[c] = "/".join(parts) if parts else get_column_letter(c)
    # 同名ヘッダーの重複には連番を付与
    seen: dict[str, int] = {}
    for c in sorted(headers):
        name = headers[c]
        if name in seen:
            seen[name] += 1
            headers[c] = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
    return headers


def pick_anchor_column(
    box: tuple[int, int, int, int], body_start: int, grid: dict[tuple[int, int], object]
) -> int:
    """レコード開始の判定に使うアンカー列（十分な密度を持つ最左列）を選ぶ。"""
    min_row, min_col, max_row, max_col = box
    counts = {
        c: sum(1 for r in range(body_start, max_row + 1) if (r, c) in grid)
        for c in range(min_col, max_col + 1)
    }
    max_count = max(counts.values(), default=0)
    threshold = max(2, max_count * 0.2)
    for c in range(min_col, max_col + 1):
        if counts[c] >= threshold:
            return c
    return min_col


def extract_marks(
    ws: Worksheet, rows: list[int], box: tuple[int, int, int, int], headers: dict[int, str]
) -> dict[str, str]:
    """レコードの行範囲内で塗りつぶされたセルを {ヘッダー名: RGB} で返す。"""
    marks: dict[str, str] = {}
    for r in rows:
        for c in range(box[1], box[3] + 1):
            fill = ws.cell(row=r, column=c).fill
            if fill.patternType is None:
                continue
            rgb = fill.fgColor.rgb if isinstance(fill.fgColor.rgb, str) else None
            if rgb in UNMARKED_RGB:
                continue
            marks.setdefault(headers[c], rgb)
    return marks


def extract_block(
    ws: Worksheet,
    box: tuple[int, int, int, int],
    grid: dict[tuple[int, int], object],
    merge_origin: dict[tuple[int, int], tuple[int, int]],
    with_fills: bool,
) -> dict:
    """1ブロックをヘッダー＋レコード列に構造化する。"""
    min_row, min_col, max_row, max_col = box
    header_start, meta = split_preamble(box, grid, merge_origin, ws.merged_cells.ranges)
    inner_box = (header_start, min_col, max_row, max_col)
    depth = detect_header_depth(inner_box, grid, merge_origin, ws.merged_cells.ranges)
    headers = composite_headers(inner_box, depth, grid)
    body_start = header_start + depth
    anchor = pick_anchor_column(inner_box, body_start, grid)

    def is_record_start(r: int) -> bool:
        """アンカー列に値があり、かつ上からの縦結合の続きでない行。"""
        if (r, anchor) not in grid:
            return False
        origin = merge_origin.get((r, anchor))
        return origin is None or origin[0] == r

    records: list[dict] = []
    current_rows: list[int] = []

    def flush() -> None:
        if not current_rows:
            return
        fields: dict[str, object] = {}
        cells: dict[str, str] = {}
        for c in range(min_col, max_col + 1):
            values, refs = [], []
            for r in current_rows:
                origin = merge_origin.get((r, c))
                key = origin or (r, c)
                if key in grid and (not refs or refs[-1] != f"{get_column_letter(key[1])}{key[0]}"):
                    ref = f"{get_column_letter(key[1])}{key[0]}"
                    if ref not in refs:
                        values.append(grid[key])
                        refs.append(ref)
            if values:
                fields[headers[c]] = values[0] if len(values) == 1 else values
                cells[headers[c]] = refs[0] if len(refs) == 1 else refs
        record = {
            "rows": [current_rows[0], current_rows[-1]],
            "fields": fields,
            "cells": cells,
        }
        # アンカー列（とその左）にしか値がない行は見出し行とみなす
        anchor_side = {headers[c] for c in range(min_col, anchor + 1)}
        if fields and set(fields).issubset(anchor_side):
            record["type"] = "section"
        if with_fills:
            marks = extract_marks(ws, list(range(current_rows[0], current_rows[-1] + 1)), box, headers)
            if marks:
                record["marks"] = marks
        records.append(record)

    for r in range(body_start, max_row + 1):
        if is_record_start(r):
            flush()
            current_rows = [r]
        elif current_rows:
            current_rows.append(r)
        elif any((r, c) in grid for c in range(min_col, max_col + 1)):
            current_rows = [r]  # アンカー値が無いまま始まる先頭行群
    flush()

    # セクション見出しを後続レコードに文脈として付与
    section: object = None
    for rec in records:
        if rec.get("type") == "section":
            section = next(iter(rec["fields"].values()))
        elif section is not None:
            rec["section"] = section

    return {
        "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
        "meta": meta,
        "header_rows": depth,
        "anchor_column": get_column_letter(anchor),
        "headers": {get_column_letter(c): h for c, h in headers.items()},
        "records": records,
    }


def write_csv(block: dict, path: Path) -> None:
    header_names = list(block["headers"].values())
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["_rows", "_type", "_section", *header_names, "_marks"])
        for rec in block["records"]:
            row = [
                f"{rec['rows'][0]}-{rec['rows'][1]}",
                rec.get("type", ""),
                rec.get("section", ""),
            ]
            for name in header_names:
                v = rec["fields"].get(name, "")
                row.append("; ".join(str(x) for x in v) if isinstance(v, list) else v)
            row.append("; ".join(f"{k}={v}" for k, v in rec.get("marks", {}).items()))
            writer.writerow(row)


def main() -> None:
    parser = argparse.ArgumentParser(description="汎用 Excel 構造化スクリプト")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--sheet", help="対象シート名（省略時は全シート）")
    parser.add_argument("--range", dest="cell_range", help="解析範囲の限定 例: A1:N260")
    parser.add_argument("--fills", action="store_true", help="塗りつぶしセルを marks として抽出")
    parser.add_argument("--list", action="store_true", help="ブロック検出結果のみ表示")
    parser.add_argument("-o", "--output", type=Path, help="JSON 出力先")
    parser.add_argument("--csv", type=Path, help="CSV 出力ディレクトリ")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    sheet_names = [args.sheet] if args.sheet else wb.sheetnames

    result = {"file": str(args.xlsx), "sheets": []}
    for name in sheet_names:
        ws = wb[name]
        grid, merge_origin = build_grid(ws)
        presence = collect_presence(ws, grid)
        if args.cell_range:
            min_c, min_r, max_c, max_r = range_boundaries(args.cell_range)
            grid = {k: v for k, v in grid.items() if min_r <= k[0] <= max_r and min_c <= k[1] <= max_c}
            presence = {k for k in presence if min_r <= k[0] <= max_r and min_c <= k[1] <= max_c}
        blocks = detect_blocks(presence, grid)

        if args.list:
            print(f"[{name}]")
            for b in blocks:
                print(f"  {get_column_letter(b[1])}{b[0]}:{get_column_letter(b[3])}{b[2]}"
                      f" ({b[2] - b[0] + 1}行 x {b[3] - b[1] + 1}列)")
            continue

        sheet_out = {"name": name, "blocks": []}
        for i, box in enumerate(blocks):
            block = extract_block(ws, box, grid, merge_origin, args.fills)
            sheet_out["blocks"].append(block)
            if args.csv:
                args.csv.mkdir(parents=True, exist_ok=True)
                safe = name.replace("/", "_")
                write_csv(block, args.csv / f"{safe}_block{i + 1}.csv")
        result["sheets"].append(sheet_out)

    if args.list:
        return
    text = json.dumps(result, ensure_ascii=False, indent=2, default=str)
    if args.output:
        args.output.write_text(text, encoding="utf-8")
        print(f"JSON: {args.output}", file=sys.stderr)
    else:
        print(text)


if __name__ == "__main__":
    main()
