"""
RAG実機能検証ハーネス（PoC検証用・既存システム非破壊）

転記後の「結果Excel」を起点に、現行RAG（Vertex AI Search の F2/F3 検索）が
実データで本当に機能するかを検証する。既存の upload→Firestore→/review フローには
一切触れず、独立してレビュー経路を回す。

検証内容（1レポートに両方出す）:
  ① 検索ヒットの中身 : 派生クエリで F2/F3(自社・他社)/補足 を直接検索し、
                       ヒットの _doc_id・struct_data・content抜粋を全件ダンプ。
  ② 最終レビュー出力 : reviewer_agent.run_review() を本番同一経路で実行し、
                       review_items + retrieval_trace を出力。

使い方:
  uv run python scripts/verify_rag.py --excel data/form_generation/output/<結果>.xlsx
  uv run python scripts/verify_rag.py --excel <...> --smoke-only      # 疎通確認のみ
  uv run python scripts/verify_rag.py --excel <...> --retrieval-only  # Geminiを呼ばず検索だけ
  オプション: --frame frameB --sheet MRC1 --utility "AA電力"

前提:
  - .env に GCP プロジェクト・Vertex AI Search データストアID が設定済み（settings.py が読込）
  - ADC（gcloud auth application-default login）または GOOGLE_APPLICATION_CREDENTIALS
  - F2/F3 が未投入なら先に scripts/create_datastores.py → scripts/ingest_knowledge.py --target all
"""
from __future__ import annotations

import argparse
import asyncio
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook

from apps.backend.app.core import settings
from apps.backend.app.core.frame_config_loader import (
    load_frame_config,
    extract_cell_definitions,
)
from apps.backend.app.agents.reviewer import knowledge_loader, reviewer_agent

REPORT_DIR = Path("data/verification")
_SNIPPET = 160  # content抜粋の最大文字数


# ── 疎通確認 ──────────────────────────────────────────────────────────────────
def smoke_check() -> dict[str, Any]:
    """設定値を表示し、F2検索を1発叩いて datastore 到達可否を判定する。"""
    info: dict[str, Any] = {
        "GCP_PROJECT_ID":     settings.GCP_PROJECT_ID,
        "GCP_LOCATION":       settings.GCP_LOCATION,
        "F2_DATASTORE_ID":    settings.VERTEX_SEARCH_F2_DATASTORE_ID,
        "F3_DATASTORE_ID":    settings.VERTEX_SEARCH_F3_DATASTORE_ID,
        "SUPPLEMENT_DS_ID":   settings.VERTEX_SEARCH_SUPPLEMENT_DATASTORE_ID or "(未設定)",
    }
    print("=" * 70)
    print(" 疎通確認 (smoke check)")
    print("=" * 70)
    for k, v in info.items():
        print(f"  {k:18}: {v or '(空)'}")

    reachable = False
    detail = ""
    try:
        hits = knowledge_loader.load_f2("NuRO", "解体", limit=1)
        reachable = True
        detail = f"load_f2('NuRO','解体',1) → {len(hits)} 件取得"
        print(f"\n  ✅ Vertex AI Search 到達OK: {detail}")
        if not hits:
            print("     ⚠️ 0件です。datastore は到達できるがデータ未投入の可能性。")
            print("        → uv run python scripts/ingest_knowledge.py --target all")
    except Exception as e:  # noqa: BLE001
        detail = f"{type(e).__name__}: {e}"
        print(f"\n  ❌ Vertex AI Search 到達NG: {detail}")
        print("     → 認証(ADC) / datastore作成 / .env を確認してください。")

    info["reachable"] = reachable
    info["smoke_detail"] = detail
    return info


# ── 結果Excel → mappings 復元 ─────────────────────────────────────────────────
def reconstruct_mappings_from_excel(
    excel_path: Path, frame: str, sheet: str
) -> list[dict]:
    """frame config の セル↔フィールド 定義を使って結果Excelから mappings を復元する。

    cell_writer の逆操作。{field_name, cell_address, value, reasoning} のリストを返す。
    同一フィールドが複数セル（計画/実績）を持つ場合は値のあるセルのみ採用する。
    """
    config = load_frame_config(frame, sheet)
    cell_defs = extract_cell_definitions(config)  # label_value / plan_actual のみ

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active

    mappings: list[dict] = []
    seen: set[tuple[str, str]] = set()

    def _add(field_name: str, cell: str) -> None:
        raw = ws[cell].value
        if raw is None or str(raw).strip() == "":
            return
        key = (field_name, cell)
        if key in seen:
            return
        seen.add(key)
        mappings.append({
            "field_name":   field_name,
            "cell_address": cell,
            "value":        str(raw).strip(),
            "reasoning":    "",
        })

    # label_value / plan_actual（既存ユーティリティが拾う分）
    for field_name, cells in cell_defs.items():
        for cell in cells:
            _add(field_name, cell)

    # tabular（解体機器表・費用内訳など）: extract_cell_definitions が拾わないため個別処理
    for section in config.get("sections", []):
        if section.get("type") != "tabular":
            continue
        _add_tabular(section, ws, _add)

    return mappings


def _add_tabular(section: dict, ws, add) -> None:
    """tabular セクションのセルを mappings に追加する。

    2形式に対応:
      ① row_match.rows で行を明示（例: MRC2 費用内訳の 人件費/材料費…）
      ② data_start_row から空行まで動的スキャン（例: MRC1 解体機器表）
    field_name は「{セクション名}_{行ID}_{列名}」。
    """
    sec_name = section.get("name", "表")
    cols = [c for c in section.get("columns", []) if c.get("column")]
    if not cols:
        return

    explicit_rows = section.get("row_match", {}).get("rows")
    if explicit_rows:
        rows = [(rm["row"], str(rm.get("row_id", rm["row"]))) for rm in explicit_rows if rm.get("row")]
    else:
        start = section.get("data_start_row")
        if not start:
            return
        end = section.get("data_end_row") or (start + 200)  # 上限ガード
        rows = []
        for r in range(int(start), int(end) + 1):
            # 全列が空になった行で打ち切り（動的表の終端検出）
            if all(ws[f"{c['column']}{r}"].value in (None, "") for c in cols):
                break
            rows.append((r, str(r)))

    for row, row_id in rows:
        for c in cols:
            add(f"{sec_name}_{row_id}_{c.get('name', c['column'])}", f"{c['column']}{row}")


def _derive_queries(mappings: list[dict]) -> dict[str, str | None]:
    """run_review と同じロジックで RAG クエリ用フィールドを取り出す。"""
    return {
        "fee_type":     reviewer_agent._extract_field(mappings, "対象費目1"),
        "reactor_type": reviewer_agent._extract_field(mappings, "炉型"),
        "utility_name": reviewer_agent._extract_field(mappings, "電力会社"),
    }


def derive_query_context(
    excel_path: Path, frame: str, target_sheet: str, context_sheet: str = "MRC1"
) -> dict[str, str | None]:
    """RAG クエリ文脈（費目・炉型・会社）を申請の基本情報シートから取得する。

    MRC2 など費目・炉型を持たないシートをレビューする場合でも、同一申請の
    基本情報シート（既定 MRC1）から費目・炉型・会社を引いてクエリに使う。
    対象シート自身に基本情報があればそちらを優先する。
    """
    q = _derive_queries(reconstruct_mappings_from_excel(excel_path, frame, target_sheet))
    if q.get("fee_type") and q.get("utility_name"):
        return q
    # 不足分を context_sheet（基本情報シート）から補完
    try:
        ctx = _derive_queries(reconstruct_mappings_from_excel(excel_path, frame, context_sheet))
    except FileNotFoundError:
        return q
    for k in ("fee_type", "reactor_type", "utility_name"):
        if not q.get(k):
            q[k] = ctx.get(k)
    return q


# ── ① 検索ヒットの中身ダンプ ──────────────────────────────────────────────────
def _format_hit(rec: dict[str, Any]) -> dict[str, Any]:
    content = rec.get("message_content") or rec.get("text_content") or ""
    struct = {k: v for k, v in rec.items() if not k.startswith("_") and k not in
              ("message_content", "text_content")}
    return {
        "doc_id":  rec.get("_doc_id", ""),
        "snippet": content[:_SNIPPET].replace("\n", " "),
        "struct":  struct,
    }


def inspect_retrieval(query_ctx: dict[str, str | None], utility_name: str) -> dict[str, Any]:
    """クエリ文脈で各Toolを直接検索し、ヒットの中身を全件ダンプする。"""
    q = query_ctx
    fee_type = q["fee_type"]

    print("\n" + "=" * 70)
    print(" ① 検索ヒットの中身 (retrieval inspection)")
    print("=" * 70)
    print(f"  派生クエリ  fee_type(対象費目1) = {fee_type!r}")
    print(f"             reactor_type(炉型)  = {q['reactor_type']!r}")
    print(f"             utility_name        = {utility_name!r}")

    tools: dict[str, list[dict]] = {
        "Tool1 F2":        knowledge_loader.load_f2("NuRO", fee_type, limit=20),
        "Tool2a F3(自社)": knowledge_loader.load_f3("NuRO", utility_name, q["reactor_type"], fee_type, None, 20),
        "Tool2b F3(他社)": knowledge_loader.load_f3("NuRO", None, q["reactor_type"], fee_type, None, 20),
        "Tool4 補足資料":  knowledge_loader.load_supplement("NuRO", utility_name, fee_type, limit=20),
    }

    result: dict[str, Any] = {"queries": q, "tools": {}}
    for name, hits in tools.items():
        formatted = [_format_hit(h) for h in hits]
        result["tools"][name] = {"count": len(hits), "hits": formatted}
        print(f"\n  ── {name}: {len(hits)} 件 ──")
        for i, h in enumerate(formatted[:10], 1):
            print(f"    [{i}] {h['doc_id']}")
            print(f"        {h['snippet']}")
            if h["struct"]:
                keys = ", ".join(f"{k}={v}" for k, v in list(h["struct"].items())[:5])
                print(f"        ({keys})")
        if len(formatted) > 10:
            print(f"    ... 他 {len(formatted) - 10} 件（レポートJSONに全件）")
    return result


# ── ② 最終レビュー出力 ───────────────────────────────────────────────────────
def run_full_review(
    mappings: list[dict], utility_name: str, frame: str, sheet: str,
    query_ctx: dict[str, str | None] | None = None,
) -> dict[str, Any]:
    """本番同一経路で run_review を実行し review_items + retrieval_trace を返す。

    query_ctx を渡すと費目・炉型を明示指定する（MRC2 など対象シートに費目が
    無い場合に、申請の基本情報シートから引いた文脈を使うため）。
    """
    print("\n" + "=" * 70)
    print(" ② 最終レビュー出力 (run_review)")
    print("=" * 70)

    query_ctx = query_ctx or {}
    review_items, trace = asyncio.run(
        reviewer_agent.run_review(
            session_id="verify-rag-harness",
            utility_name=utility_name,
            mappings=mappings,
            frame_name=frame,
            sheet_name=sheet,
            reactor_type=query_ctx.get("reactor_type"),
            fee_type=query_ctx.get("fee_type"),
        )
    )

    items = [i.model_dump() for i in review_items]
    print(f"  指摘 {len(items)} 件")
    for it in items:
        print(f"\n  ● [{it['severity']}] {it['field_name']} ({it['cell_address']})"
              f"  src={it['knowledge_source']}")
        print(f"    {it['comment']}")
        if it.get("evidence"):
            print(f"    └ 根拠: {str(it['evidence'])[:_SNIPPET]}")

    print("\n  -- retrieval_trace --")
    for t in trace:
        print(f"    {t.get('tool','')}: {t.get('count',0)}件 query={t.get('query','')!r}")

    return {"review_items": items, "retrieval_trace": trace}


# ── レポート書き出し ─────────────────────────────────────────────────────────
def write_report(excel_path: Path, payload: dict[str, Any]) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = excel_path.stem
    json_path = REPORT_DIR / f"{stem}_{ts}.json"
    md_path = REPORT_DIR / f"{stem}_{ts}.md"

    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [f"# RAG検証レポート: {stem}", "", f"- 実行: {ts}", f"- 入力Excel: `{excel_path}`", ""]
    smoke = payload.get("smoke", {})
    lines += ["## 疎通", f"- 到達: {'OK' if smoke.get('reachable') else 'NG'} ({smoke.get('smoke_detail','')})",
              f"- F2 datastore: `{smoke.get('F2_DATASTORE_ID','')}` / F3: `{smoke.get('F3_DATASTORE_ID','')}`", ""]

    q = payload.get("retrieval", {}).get("queries", {})
    lines += ["## 復元 mappings / 派生クエリ",
              f"- mappings 件数: {len(payload.get('mappings', []))}",
              f"- fee_type(対象費目1): `{q.get('fee_type')}` / 炉型: `{q.get('reactor_type')}`"
              f" / 会社: `{q.get('utility_name')}`", ""]

    ret = payload.get("retrieval", {}).get("tools", {})
    lines += ["## ① 検索ヒット件数"]
    for name, d in ret.items():
        lines.append(f"- {name}: **{d['count']}** 件")
    lines.append("")

    rv = payload.get("review", {})
    if rv:
        items = rv.get("review_items", [])
        lines += ["## ② レビュー指摘", f"- 合計 {len(items)} 件"]
        for it in items:
            lines.append(f"  - [{it['severity']}] {it['field_name']} ({it['cell_address']}) "
                         f"src={it['knowledge_source']}: {it['comment']}")
        lines.append("")
    lines += ["> 検索ヒットの中身（全件）と struct_data は同名 `.json` を参照。"]

    md_path.write_text("\n".join(lines), encoding="utf-8")
    return md_path


# ── エントリーポイント ───────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(description="RAG実機能検証ハーネス")
    parser.add_argument("--excel", help="転記結果Excelのパス")
    parser.add_argument("--frame", default="frameB")
    parser.add_argument("--sheet", default="MRC1")
    parser.add_argument("--context-sheet", default="MRC1",
                        help="費目・炉型・会社のクエリ文脈を取る基本情報シート（既定MRC1）")
    parser.add_argument("--utility", default=None, help="電力会社名（未指定ならExcelから取得）")
    parser.add_argument("--smoke-only", action="store_true", help="疎通確認のみ")
    parser.add_argument("--retrieval-only", action="store_true", help="Geminiを呼ばず検索だけ")
    args = parser.parse_args()

    payload: dict[str, Any] = {}
    payload["smoke"] = smoke_check()

    if args.smoke_only:
        return

    if not args.excel:
        print("\n--excel を指定してください（--smoke-only 以外）")
        sys.exit(1)
    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"\nExcelが見つかりません: {excel_path}")
        sys.exit(1)

    mappings = reconstruct_mappings_from_excel(excel_path, args.frame, args.sheet)
    payload["mappings"] = mappings

    # RAGクエリ文脈（費目・炉型・会社）は申請の基本情報シート(MRC1)から取得。
    # MRC2 など対象シートに費目が無くても申請単位でクエリを組み立てられる。
    query_ctx = derive_query_context(excel_path, args.frame, args.sheet, args.context_sheet)
    utility = args.utility or query_ctx.get("utility_name") or "不明電力"
    print(f"\n復元した mappings: {len(mappings)} 件 / 電力会社={utility}")

    payload["retrieval"] = inspect_retrieval(query_ctx, utility)

    if not args.retrieval_only:
        payload["review"] = run_full_review(mappings, utility, args.frame, args.sheet, query_ctx)

    report = write_report(excel_path, payload)
    print(f"\n📄 レポート出力: {report} (+ .json)")


if __name__ == "__main__":
    main()
