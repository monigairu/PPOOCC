"""Excel抽出の正当性検証ハーネス（生Excelとの全件突合・既存システム非破壊）

「Excelから正しく値が取れているか」を、抽出コードとは独立の実装で
生Excelを読み直して全件突合する。検証対象は2系統:

  ① ナレッジ供給 : F2/F3ナレッジExcel → _read_excel_by_schema()（平坦化）の突合。
                   行の取りこぼし・message_id採番ズレ・内容不一致・
                   空セルへの上行値の染み出し（ffill捏造）を検出する。
                   --bigquery を付けると BigQuery 実テーブルとも突合する。
  ② レビュー様式 : 転記結果Excel → reconstruct_mappings_from_excel() の突合。
                   configが定義する全セル（label_value/plan_actual/tabular）の
                   取り逃し・値違い・動的表の空行打ち切りによる取り逃しを検出する。

使い方:
  uv run python scripts/preliminary_review/verify_extraction.py                        # ①ナレッジのみ
  uv run python scripts/preliminary_review/verify_extraction.py --bigquery             # ①＋BigQuery突合
  uv run python scripts/preliminary_review/verify_extraction.py --review-excel <転記結果.xlsx> [--frame frameB]
  （組み合わせ可。終了コード: 問題なし=0 / 問題あり=1）

前提:
  - ①は data/knowledge/ のExcelと schema/ のYAMLのみ（GCP不要）
  - --bigquery は .env のGCP設定とADC（BigQuery読み取り権限）が必要
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from openpyxl import load_workbook

from apps.backend.app.preliminary_review.knowledge.excel_reader import (
    _col_letter_to_idx,
    _discover_schemas,
    _infer_direction,
    _read_excel_by_schema,
    excel_to_bq_input,
    read_all_f2,
    read_all_f3,
)

KNOWLEDGE_DIR = Path("data/knowledge")

# 検出した問題（最後にまとめて表示・1件でもあれば終了コード1）
problems: list[str] = []
# 問題ではない補足情報（メッセージ0件のIDなど）
notes: list[str] = []


def build_raw_grid(ws) -> dict[tuple[int, int], str]:
    """ワークシートの生セル値を、結合セルのみアンカー値で展開して返す。

    抽出コード（_read_excel_by_schema）とは独立に「Excelの正しい姿」を
    作るための基準データ。結合セルはExcel仕様上アンカー（左上）にしか
    値を持たないため、結合範囲全体へ展開する。それ以外の空セルは空のまま
    （＝空セルに値が入っていたら抽出側の捏造と判定できる）。

    Args:
        ws: openpyxl のワークシート（data_only=True で読み込んだもの）。

    Returns:
        {(行番号, 列番号): セル値文字列} の辞書（1始まり・空セルは ""）。
    """
    grid: dict[tuple[int, int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            grid[(cell.row, cell.column)] = "" if v is None else str(v).strip()
    for rng in ws.merged_cells.ranges:
        anchor = grid.get((rng.min_row, rng.min_col), "")
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                grid[(r, c)] = anchor
    return grid


def expected_records_from_raw(schema: dict, file_path: Path) -> dict[str, dict]:
    """生Excelから「期待される平坦化結果」をスキーマ定義に沿って独立に組み立てる。

    _read_excel_by_schema() と同じスキーマ（fixed_columns / repeating_qa_columns）を
    参照するが、読み取りは build_raw_grid() の生グリッドから行うため、
    抽出コードのバグ（前埋め・行スキップ等）に影響されない。

    Args:
        schema: data/knowledge/schema/ のYAML（f2_*/f3_*_schema.yaml）の内容。
        file_path: 対象ナレッジExcelのパス。

    Returns:
        ID → 期待値 の辞書。各値は
            {"row": Excel行番号,
             "fixed": {列キー: 値} （固定列の期待値）,
             "messages": [(message_id, round, direction, 内容), ...]（読み順）}。
    """
    wb = load_workbook(file_path, data_only=True)
    sheet = schema.get("excel_sheet")
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    grid = build_raw_grid(ws)

    start = schema.get("layout", {}).get("data_start_row", 7)
    id_col = _col_letter_to_idx(schema.get("loader_config", {}).get("id_column", "A")) + 1
    fixed = schema.get("fixed_columns", [])
    qa = schema.get("repeating_qa_columns")

    expected: dict[str, dict] = {}
    for r in range(start, ws.max_row + 1):
        id_val = grid.get((r, id_col), "")
        if not id_val or id_val in ("nan", "None"):
            continue
        if id_val in expected:
            problems.append(f"[{file_path.name}:{sheet}] 行{r}: ID重複 {id_val}（複数行に同一ID）")
            continue
        fixed_vals = {
            cd["key"]: grid.get((r, _col_letter_to_idx(cd["col"]) + 1), "") for cd in fixed
        }
        msgs = []
        if qa:
            start_c = _col_letter_to_idx(qa["start_col"]) + 1
            seq = 0
            for rnd in range(1, qa["max_rounds"] + 1):
                for fd in qa["fields"]:
                    c = start_c + (rnd - 1) * qa["col_per_round"] + fd["col_offset"]
                    content = grid.get((r, c), "")
                    if not content or content in ("nan", "None"):
                        continue
                    seq += 1
                    msgs.append((f"{id_val}_{seq:02d}", rnd, _infer_direction(fd["key"]), content))
        expected[id_val] = {"row": r, "fixed": fixed_vals, "messages": msgs}
    return expected


def check_knowledge(prefix: str) -> None:
    """①ナレッジ供給の突合: 生Excel vs _read_excel_by_schema() を全件比較する。

    検出する問題（problems に追記）:
      - 行ごと消失（生ExcelにあるIDが出力に無い）
      - 捏造（生Excelに無いID/メッセージが出力にある＝旧ffillバグの再発）
      - message_id の採番ズレ・round/direction/内容の不一致
      - 固定列の値違い（空セルへの上行値の染み出しを含む）

    Args:
        prefix: "f2" または "f3"（schema/ の {prefix}_*_schema.yaml を対象にする）。
    """
    print(f"\n{'=' * 72}\n ① ナレッジ {prefix.upper()}: 生Excel vs 平坦化抽出\n{'=' * 72}")
    total_exp = total_act = 0
    for schema in _discover_schemas(prefix):
        sname = schema.get("sheet_name", "")
        excel_file = schema.get("excel_file") or f"{schema.get('frame', '').upper()}_{sname}.xlsx"
        fp = KNOWLEDGE_DIR / excel_file
        if not fp.exists():
            notes.append(f"[{prefix}] {excel_file} なし（スキップ）")
            continue

        expected = expected_records_from_raw(schema, fp)
        actual_records, _ = _read_excel_by_schema(schema, fp)
        actual: dict[str, list] = {}
        for rec in actual_records:
            actual.setdefault(rec["id"], []).append(rec)

        n_exp = sum(len(v["messages"]) for v in expected.values())
        total_exp += n_exp
        total_act += len(actual_records)
        tag = f"{excel_file}:{sname}"
        status = "OK" if n_exp == len(actual_records) else "❌"
        print(f"  {status} {tag:44} 期待 {len(expected):3}ID/{n_exp:4}msg"
              f"  実際 {len(actual):3}ID/{len(actual_records):4}msg")

        for missing in sorted(set(expected) - set(actual)):
            if expected[missing]["messages"]:
                problems.append(f"[{tag}] ID {missing}: 出力に存在しない（行ごと消失）")
            else:
                notes.append(f"[{tag}] ID {missing}: メッセージ0件のためレコード化されず（仕様どおり）")
        for extra in sorted(set(actual) - set(expected)):
            problems.append(f"[{tag}] ID {extra}: 生Excelに無いIDが出力された（捏造の疑い）")

        for id_val in sorted(set(expected) & set(actual)):
            exp_msgs = expected[id_val]["messages"]
            act_msgs = sorted(actual[id_val], key=lambda x: x["message_id"])
            if len(exp_msgs) != len(act_msgs):
                problems.append(f"[{tag}] ID {id_val}: メッセージ数不一致 "
                                f"期待{len(exp_msgs)} 実際{len(act_msgs)}（捏造/消失）")
                continue
            for (mid, rnd, dirn, content), rec in zip(exp_msgs, act_msgs):
                if rec["message_id"] != mid:
                    problems.append(f"[{tag}] {id_val}: message_id ずれ 期待{mid} 実際{rec['message_id']}")
                if rec["message_content"] != content:
                    problems.append(f"[{tag}] {mid}: 内容不一致 期待={content[:50]!r} 実際={rec['message_content'][:50]!r}")
                if rec.get("round") != rnd:
                    problems.append(f"[{tag}] {mid}: round 期待{rnd} 実際{rec.get('round')}")
                if rec.get("message_direction") != dirn:
                    problems.append(f"[{tag}] {mid}: direction 期待{dirn} 実際{rec.get('message_direction')}")
            for k, ev in expected[id_val]["fixed"].items():
                av = act_msgs[0].get(k, "")
                if str(av) != str(ev):
                    hint = "（空セルへの上行値の染み出し）" if not ev else ""
                    problems.append(f"[{tag}] ID {id_val} 固定列 {k}: 期待 {ev!r} 実際 {av!r}{hint}")
    print(f"     {'合計':44} 期待メッセージ {total_exp} / 実際 {total_act}")


def check_bigquery() -> None:
    """BigQuery実テーブルの突合: Excelから再生成した行と message_id・内容を比較する。

    「Excel（正本）→ excel_to_bq_input()」で今この場で作った行リストと、
    BigQuery 上の f2/f3 平坦テーブルの全行を突合し、投入漏れ・余剰行
    （過去バグの残骸等）・内容の食い違いを検出する。

    前提: .env の GCP 設定と ADC（BigQuery 読み取り権限）。
    """
    print(f"\n{'=' * 72}\n ① BigQuery 実テーブル vs Excel再生成\n{'=' * 72}")
    from google.cloud import bigquery

    from apps.backend.app.core import settings
    from apps.backend.app.preliminary_review import config as review_config

    client = bigquery.Client(project=settings.GCP_PROJECT_ID)
    for ktype, reader, table in (
        ("f2", read_all_f2, review_config.BIGQUERY_F2_TABLE_ID),
        ("f3", read_all_f3, review_config.BIGQUERY_F3_TABLE_ID),
    ):
        rows = excel_to_bq_input(reader(), ktype)
        local = {r["message_id"]: r["message_content"] for r in rows}
        if len(local) != len(rows):
            problems.append(f"[BQ:{ktype}] ローカル行に message_id 重複 {len(rows) - len(local)} 件")

        fq = f"{settings.GCP_PROJECT_ID}.{review_config.BIGQUERY_DATASET_ID}.{table}"
        bq_rows = {
            r["message_id"]: r["message_content"]
            for r in client.query(f"SELECT message_id, message_content FROM `{fq}`").result()
        }

        only_local = set(local) - set(bq_rows)
        only_bq = set(bq_rows) - set(local)
        diff = [m for m in set(local) & set(bq_rows) if local[m] != bq_rows[m]]
        status = "OK" if not (only_local or only_bq or diff) else "❌"
        print(f"  {status} {ktype.upper()}: Excel再生成 {len(rows)} 行 / BQ {len(bq_rows)} 行"
              f" / 投入漏れ {len(only_local)} / BQ余剰 {len(only_bq)} / 内容差分 {len(diff)}")
        for mid in sorted(only_local)[:10]:
            problems.append(f"[BQ:{ktype}] BQ に無い message_id（投入漏れ）: {mid}")
        for mid in sorted(only_bq)[:10]:
            problems.append(f"[BQ:{ktype}] Excel由来に無い message_id が BQ にある（要再ingest）: {mid}")
        for mid in diff[:10]:
            problems.append(f"[BQ:{ktype}] {mid}: 内容が BQ と不一致（要再ingest）")


def check_review_form(excel_path: Path, frame: str) -> None:
    """②レビュー様式の突合: 生Excel vs reconstruct_mappings_from_excel() を比較する。

    frame の全シートについて、configが定義する全セル（label_value/plan_actual/
    tabular）を生セル値と突き合わせる。検出する問題:
      - 取り逃し（生Excelに値がある定義セルが mappings に無い）
      - 値違い（mappings の値が生セル値と異なる＝でっち上げ含む）
      - 動的表の打ち切り（空行の後に値が再開しているのにスキャンが止まった）

    Args:
        excel_path: 転記結果Excelのパス。
        frame: 様式名（config/{frame}/ のシート定義を参照）。
    """
    print(f"\n{'=' * 72}\n ② レビュー様式: 生Excel vs mappings復元（{excel_path.name}）\n{'=' * 72}")
    from apps.backend.app.preliminary_review.knowledge.result_reader import reconstruct_mappings_from_excel
    from apps.backend.app.core.frame_config_loader import (
        extract_cell_definitions,
        list_frame_sheets,
        load_frame_config,
    )

    wb = load_workbook(excel_path, data_only=True)
    for sheet in list_frame_sheets(frame):
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        grid = build_raw_grid(ws)
        config = load_frame_config(frame, sheet)
        mappings = reconstruct_mappings_from_excel(excel_path, frame, sheet)
        mapped = {m["cell_address"]: m["value"] for m in mappings}

        # configが定義する全セルを列挙（cell -> field_name）
        defined: dict[str, str] = {}
        for field, cells in extract_cell_definitions(config).items():
            for cell in cells:
                defined[cell] = field
        gap_rows: list[tuple[str, int]] = []
        for section in config.get("sections", []):
            if section.get("type") != "tabular":
                continue
            cols = [c for c in section.get("columns", []) if c.get("column")]
            explicit = section.get("row_match", {}).get("rows")
            if explicit:
                rows = [rm["row"] for rm in explicit if rm.get("row")]
            else:
                s = section.get("data_start_row")
                if not s:
                    continue
                rows = list(range(int(s), int(section.get("data_end_row") or s + 200) + 1))
            blank_seen = False
            for r in rows:
                vals = [grid.get((r, _col_letter_to_idx(c["column"]) + 1), "") for c in cols]
                if not explicit:
                    if all(v == "" for v in vals):
                        blank_seen = True
                        continue
                    if blank_seen:
                        gap_rows.append((section.get("name", "表"), r))
                for c in cols:
                    defined[f"{c['column']}{r}"] = f"{section.get('name', '表')}_{r}_{c.get('name', c['column'])}"

        def _raw(cell: str) -> str:
            col = "".join(ch for ch in cell if ch.isalpha())
            row = int("".join(ch for ch in cell if ch.isdigit()))
            return grid.get((row, _col_letter_to_idx(col) + 1), "")

        missed = [(c, f) for c, f in defined.items() if _raw(c) != "" and c not in mapped]
        mismatched = [(c, f) for c, f in defined.items()
                      if _raw(c) != "" and c in mapped and mapped[c] != _raw(c)]
        phantom = [m for m in mappings if _raw(m["cell_address"]) != m["value"]]

        status = "OK" if not (missed or mismatched or phantom or gap_rows) else "❌"
        print(f"  {status} {sheet}: 定義セル {len(defined)} / mappings {len(mappings)}"
              f" / 取り逃し {len(missed)} / 値違い {len(mismatched) + len(phantom)}"
              f" / 動的表の打ち切り {len(gap_rows)}")
        for cell, field in missed:
            problems.append(f"[様式:{sheet}] {cell}({field}): 値 {_raw(cell)[:40]!r} が mappings に無い（取り逃し）")
        for cell, field in mismatched:
            problems.append(f"[様式:{sheet}] {cell}({field}): 期待 {_raw(cell)[:40]!r} 実際 {mapped[cell][:40]!r}")
        for m in phantom:
            problems.append(f"[様式:{sheet}] {m['cell_address']}({m['field_name']}): "
                            f"mappings値 {m['value'][:40]!r} だが生セルは {_raw(m['cell_address'])[:40]!r}")
        for name, r in gap_rows:
            problems.append(f"[様式:{sheet}] 表{name} 行{r}: 空行の後に値がある（動的スキャン打ち切りで取り逃し）")


def dump_knowledge(prefix: str, limit_ids: int) -> None:
    """ナレッジExcelの読み込み結果を人間が読める形で表示する（突合ではなく可視化）。

    シートごとに、ID（=1つの問合せスレッド＝Excelの1行）単位で
    「固定列（メタデータ）→ メッセージの通し連番チャンク」の順に表示する。
    Excelの横持ち（1行にやり取りが横に連なる）が、どう縦持ち
    （1メッセージ=1行・message_id付き）に平坦化されたかを目で確認できる。

    Args:
        prefix: "f2" または "f3"。
        limit_ids: シートごとに表示するID数の上限（0以下なら全件）。
    """
    print(f"\n{'=' * 72}\n 📖 ナレッジ {prefix.upper()}: Excel読み込み結果ダンプ\n{'=' * 72}")
    for schema in _discover_schemas(prefix):
        sname = schema.get("sheet_name", "")
        excel_file = schema.get("excel_file") or f"{schema.get('frame', '').upper()}_{sname}.xlsx"
        fp = KNOWLEDGE_DIR / excel_file
        if not fp.exists():
            continue
        records, utility = _read_excel_by_schema(schema, fp)

        # 出現順を保って ID ごとにまとめる
        by_id: dict[str, list] = {}
        for rec in records:
            by_id.setdefault(rec["id"], []).append(rec)

        print(f"\n─── {excel_file} [{sname}] "
              f"（{len(by_id)}スレッド / {len(records)}メッセージ"
              f"{f' / 電力会社={utility}' if utility else ''}）───")
        shown = list(by_id.items())[:limit_ids] if limit_ids > 0 else list(by_id.items())
        for id_val, msgs in shown:
            first = msgs[0]
            # メッセージ系以外の固定列（メタデータ）を1行に要約
            meta = {k: v for k, v in first.items()
                    if k not in ("message_id", "message_content", "message_direction", "round")
                    and v not in ("", None)}
            print(f"\n  ● ID {id_val}  （Excel 1行 → {len(msgs)}メッセージに平坦化）")
            for k, v in meta.items():
                print(f"      {k:20}: {str(v)[:60]}")
            for m in msgs:
                arrow = "→NuRO確認" if m["message_direction"] == "nuro" else "←電力回答"
                print(f"      [{m['message_id']}] r{m['round']} {arrow}: {m['message_content'][:56]}")
        if limit_ids > 0 and len(by_id) > limit_ids:
            print(f"\n  … 他 {len(by_id) - limit_ids} スレッド（--dump 0 で全件表示）")


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel抽出の正当性検証（生Excelとの全件突合）")
    parser.add_argument("--bigquery", action="store_true",
                        help="BigQuery実テーブルとも突合する（要GCP認証）")
    parser.add_argument("--review-excel", default=None,
                        help="転記結果Excelのパス（②レビュー様式の突合を実行）")
    parser.add_argument("--frame", default="frameB", help="②で使う様式名（既定 frameB）")
    parser.add_argument("--dump", nargs="?", type=int, const=2, default=None, metavar="N",
                        help="読み込み結果をIDごとに表示（N=シートごとの表示スレッド数・省略時2・0で全件）。突合は行わない")
    args = parser.parse_args()

    if args.dump is not None:
        dump_knowledge("f2", args.dump)
        dump_knowledge("f3", args.dump)
        return

    check_knowledge("f2")
    check_knowledge("f3")
    if args.bigquery:
        check_bigquery()
    if args.review_excel:
        excel_path = Path(args.review_excel)
        if not excel_path.exists():
            print(f"転記結果Excelが見つかりません: {excel_path}")
            sys.exit(1)
        check_review_form(excel_path, args.frame)

    print(f"\n{'=' * 72}\n 結果\n{'=' * 72}")
    if problems:
        print(f"❌ 問題 {len(problems)} 件:")
        for p in problems:
            print(f"  - {p}")
        sys.exit(1)
    print("✅ 問題なし（生Excelと抽出結果は完全一致）")
    if notes:
        print(f"\nℹ️ 補足 {len(notes)} 件:")
        for n in notes:
            print(f"  - {n}")


if __name__ == "__main__":
    main()
