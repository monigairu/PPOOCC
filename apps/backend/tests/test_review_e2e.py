"""
レビュー機能 E2E テスト

テスト対象:
  1. detect_plan_diff       — 計画・実績差分検出ロジック
  2. knowledge_loader       — フィルタリング・権限制御・ファイルなし時の安全動作
  3. POST /api/review       — Firestore + Gemini をモックして全体フローを検証
  4. POST /api/review/{id}/feedback — 承諾・棄却の Firestore 保存制御
  5. GET  /api/review/sessions     — 未レビューセッション一覧

モック方針:
  - Firestore: unittest.mock.MagicMock（接続なし）
  - Gemini:    unittest.mock.patch("apps.backend.app.core.ai_client.call_gemini")
  - Excel:     pandas DataFrame を直接返すように knowledge_loader をパッチ
"""
import json
from unittest.mock import MagicMock, patch, AsyncMock
import pytest
from fastapi.testclient import TestClient

from apps.backend.app.preliminary_review.agent import detect_plan_diff, _evaluate_diff, _to_number


# ─────────────────────────────────────────────────────────────────────────────
# 1. detect_plan_diff ユニットテスト
# ─────────────────────────────────────────────────────────────────────────────

class TestDetectPlanDiff:
    """計画・実績差分検出（設定駆動）

    detect_plan_diff は frame config の plan_actual セクションを読み、
    plan セル(G列)と actual セル(K列)の両方を持つフィールドのみ比較する。
    本テストは「計画/実績の両方を持つフィールド」を使ってロジックを検証する。
    現行 MRC1.yaml で plan+actual を両方持つ代表フィールドは「工期開始日」(G10/K10)。
    総額(G18/K18) は MRC2 の SUM 数式へ移管され plan_actual から外れたため、
    MRC1 では差分検出の対象外（その仕様は test_sougaku_moved_to_mrc2 で固定）。
    """

    # 工期開始日 = 現行 MRC1.yaml で {plan: G10, actual: K10} を持つ代表フィールド
    _FIELD = "工期開始日"
    _PLAN_CELL = "G10"
    _ACTUAL_CELL = "K10"

    def _make_mappings(self, kubun: str, plan: str = "", actual: str = "") -> list[dict]:
        mappings = [{"field_name": "計画実績区分", "cell_address": "C8", "value": kubun, "reasoning": ""}]
        if plan:
            mappings.append({"field_name": self._FIELD, "cell_address": self._PLAN_CELL, "value": plan, "reasoning": ""})
        if actual:
            mappings.append({"field_name": self._FIELD, "cell_address": self._ACTUAL_CELL, "value": actual, "reasoning": ""})
        return mappings

    def test_keikaku_returns_empty(self):
        """計画提出は差分チェックをスキップして空リストを返す"""
        diffs = detect_plan_diff(self._make_mappings("計画", "1000", "1500"))
        assert diffs == []

    def test_jisseki_detects_large_diff(self):
        """実績提出で10%超の乖離を検出する"""
        diffs = detect_plan_diff(self._make_mappings("実績", "1000", "1300"))
        assert len(diffs) == 1
        assert diffs[0]["field_name"] == self._FIELD
        assert "30.0%" in diffs[0]["diff_note"]

    def test_jisseki_ignores_small_diff(self):
        """実績提出でも10%未満の差異は無視する"""
        diffs = detect_plan_diff(self._make_mappings("実績", "1000", "1050"))
        assert diffs == []

    def test_jisseki_detects_missing_plan(self):
        """計画値が未記入で実績値のみある場合を検出する"""
        diffs = detect_plan_diff(self._make_mappings("実績", "", "1500"))
        assert len(diffs) == 1
        assert "計画値が未記入" in diffs[0]["diff_note"]

    def test_jisseki_detects_missing_actual(self):
        """実績値が未記入で計画値のみある場合を検出する"""
        diffs = detect_plan_diff(self._make_mappings("実績", "1000", ""))
        assert len(diffs) == 1
        assert "実績値が未記入" in diffs[0]["diff_note"]

    def test_both_empty_no_diff(self):
        """計画・実績ともに空のフィールドは差分なしとして無視する"""
        diffs = detect_plan_diff(self._make_mappings("実績", "", ""))
        assert diffs == []

    def test_kubun_missing_returns_empty(self):
        """計画実績区分が mappings にない場合は安全に空リストを返す"""
        diffs = detect_plan_diff([
            {"field_name": self._FIELD, "cell_address": self._PLAN_CELL, "value": "1000", "reasoning": ""},
            {"field_name": self._FIELD, "cell_address": self._ACTUAL_CELL, "value": "9999", "reasoning": ""},
        ])
        assert diffs == []

    def test_sougaku_moved_to_mrc2(self):
        """総額(G18/K18)は MRC2 の SUM 数式へ移管され、MRC1 の plan_actual から除外された。

        設計上の意思決定を固定するテスト。総額に大きな乖離があっても MRC1 では
        差分検出されない（数値の計画/実績比較は MRC2 側の責務へ移る）。
        ※ MRC2.yaml に計画/実績の数値ペアが定義された時点で detect_plan_diff(sheet="MRC2")
          が設定駆動で自動的に対象化する（コード変更不要）。
        """
        mappings = [
            {"field_name": "計画実績区分", "cell_address": "C8",  "value": "実績", "reasoning": ""},
            {"field_name": "総額",         "cell_address": "G18", "value": "1000", "reasoning": ""},
            {"field_name": "総額",         "cell_address": "K18", "value": "9999", "reasoning": ""},
        ]
        assert detect_plan_diff(mappings, sheet_name="MRC1") == []


class TestEvaluateDiff:
    """差分評価ロジックのユニットテスト"""

    def test_equal_values_returns_none(self):
        assert _evaluate_diff("1000", "1000") is None

    def test_numeric_below_threshold(self):
        assert _evaluate_diff("1000", "1090") is None  # 9% < 10%

    def test_numeric_above_threshold(self):
        result = _evaluate_diff("1000", "1100")
        assert result is not None
        assert "10.0%" in result

    def test_comma_number(self):
        """カンマ区切り数値を正しく解釈する"""
        result = _evaluate_diff("1,000", "2,000")
        assert result is not None  # 100% 差

    def test_string_diff(self):
        """文字列の差異を検出する"""
        result = _evaluate_diff("計画A", "実績B")
        assert result is not None


class TestToNumber:
    def test_plain(self):
        assert _to_number("1000") == 1000.0

    def test_comma(self):
        assert _to_number("1,000") == 1000.0

    def test_unit(self):
        assert _to_number("1000千円") == 1000.0

    def test_invalid(self):
        assert _to_number("N/A") is None

    def test_empty(self):
        assert _to_number("") is None


# ─────────────────────────────────────────────────────────────────────────────
# 2. knowledge_loader ユニットテスト
# ─────────────────────────────────────────────────────────────────────────────

class TestKnowledgeLoader:
    """Excelファイルなしでも安全に動作することを確認"""

    def test_f2_returns_empty_for_denryoku(self):
        """電力ロールは F2 ナレッジを参照できない"""
        from apps.backend.app.preliminary_review.knowledge.knowledge_loader import load_f2
        result = load_f2(caller_role="電力")
        assert result == []

    def test_f2_returns_empty_list_when_no_file(self):
        """F2ファイルが存在しない場合は空リストを返す（エラーにならない）"""
        from apps.backend.app.preliminary_review.knowledge.knowledge_loader import load_f2
        result = load_f2(caller_role="NuRO")
        assert isinstance(result, list)

    def test_f3_returns_empty_list_when_no_file(self):
        """F3ファイルが存在しない場合は空リストを返す（エラーにならない）"""
        from apps.backend.app.preliminary_review.knowledge.knowledge_loader import load_f3
        result = load_f3(caller_role="NuRO", utility_name=None)
        assert isinstance(result, list)

    def test_f3_denryoku_requires_utility_name(self):
        """電力ロールで utility_name なしは空リストを返す"""
        from apps.backend.app.preliminary_review.knowledge.knowledge_loader import load_f3
        result = load_f3(caller_role="電力", utility_name=None)
        assert result == []

    def test_schema_discovery(self):
        """スキーマファイルが正しく検出される（Phase2: _excel_reader に移動）

        F3 は北の海電力(F3_knowledge.xlsx)4枚に加え、関東電力PoC評価用
        (F3_knowledge_関東電力.xlsx)4枚を追加したため計8枚。両系統が検出されることを確認する。
        """
        from apps.backend.app.preliminary_review.knowledge.excel_reader import _discover_schemas
        f3 = _discover_schemas("f3")
        f2 = _discover_schemas("f2")
        f3_files = {s.get("excel_file") for s in f3}
        assert "F3_knowledge.xlsx" in f3_files, f"北の海電力F3が未検出: {f3_files}"
        assert "F3_knowledge_関東電力.xlsx" in f3_files, f"関東F3が未検出: {f3_files}"
        assert len(f3) >= 4, f"F3スキーマは4件以上のはず: {[s['sheet_name'] for s in f3]}"
        assert len(f2) == 2, f"F2スキーマは2件のはず: {[s['sheet_name'] for s in f2]}"

    def test_f3_schema_sheet_names(self):
        """F3スキーマのsheet_nameが正しい（Phase2: _excel_reader に移動）"""
        from apps.backend.app.preliminary_review.knowledge.excel_reader import _discover_schemas
        sheets = {s["sheet_name"] for s in _discover_schemas("f3")}
        assert sheets == {"KNI_1G_01", "KNI_1G_02", "KNI_1G_03", "KNI_2G"}

    def test_excel_reader_with_mock_data(self, tmp_path):
        """小さな実Excelファイルで読み込み処理を検証する（Phase2: _excel_reader に移動）"""
        from openpyxl import Workbook
        from apps.backend.app.preliminary_review.knowledge.excel_reader import _read_excel_by_schema

        schema = {
            "layout": {"data_start_row": 1},
            "loader_config": {"id_column": "A"},
            "fixed_columns": [
                {"key": "id",            "col": "A", "dtype": "string"},
                {"key": "cost_category", "col": "B", "dtype": "string"},
            ],
            "repeating_qa_columns": {
                "start_col": "C", "col_per_round": 2, "max_rounds": 1,
                "fields": [
                    {"key": "nuro_comment",  "col_offset": 0},
                    {"key": "denryoku_reply","col_offset": 1},
                ],
            },
            "output_model": {"flatten_qa": True},
            "meta_cells": {},
        }

        wb = Workbook()
        wb.active.append(["03_1G_01_0001", "維持管理費", "確認します", "問題ありません"])
        xlsx = tmp_path / "mini_f3.xlsx"
        wb.save(xlsx)

        records, utility = _read_excel_by_schema(schema, xlsx)

        assert utility == ""
        assert len(records) == 2
        assert records[0]["id"] == "03_1G_01_0001"
        assert records[0]["message_direction"] == "nuro"
        assert records[1]["message_direction"] == "denryoku"
        # message_id は公式 ver5.3 準拠の通し連番（読み順で 01, 02, …）
        assert records[0]["message_id"] == "03_1G_01_0001_01"
        assert records[1]["message_id"] == "03_1G_01_0001_02"
        assert records[0]["round"] == 1 and records[1]["round"] == 1

    def test_derive_reactor_type_map_lookup(self):
        """炉型導出：発電所キー＋号機上書きキーの優先順位（2026-07-02設計確定）"""
        import apps.backend.app.preliminary_review.knowledge.excel_reader as xr
        original = xr._plant_reactor_map
        xr._plant_reactor_map = {"敦賀発電所": "BWR", "敦賀発電所/2号機": "PWR"}
        try:
            assert xr.derive_reactor_type("敦賀発電所", "1号機") == "BWR"   # 発電所キー
            assert xr.derive_reactor_type("敦賀発電所", "2号機") == "PWR"   # 号機上書きが優先
            assert xr.derive_reactor_type("不明発電所") == ""               # マップ外は不明
            assert xr.derive_reactor_type("") == ""
        finally:
            xr._plant_reactor_map = original

    def test_f3_reactor_type_derived_and_consistent(self):
        """炉型は該当発電所から導出され、発電所・号機単位で一貫する（2026-07-02設計確定）

        正本Excel（ver5.3様式）に炉型の列は存在しない（電力会社が編集する様式に
        列を足さない）。plant_reactor_map.yaml のドメイン知識から導出するため、
        同一発電所・号機に複数の炉型が混ざることは構造的にありえない。
        """
        from apps.backend.app.preliminary_review.knowledge.excel_reader import read_all_f3
        records = read_all_f3()
        with_plant = [r for r in records if r.get("plant_site")]
        assert with_plant, "該当発電所を持つF3レコードが読めない"

        derived = {r.get("reactor_type", "") for r in with_plant} - {""}
        assert derived, "炉型が1件も導出されない（plant_reactor_map.yaml の発電所名と不一致？）"

        seen: dict = {}
        for r in with_plant:
            key = (r["plant_site"], r.get("plant_unit", ""))
            rt = r.get("reactor_type", "")
            assert seen.setdefault(key, rt) == rt, f"同一発電所・号機で炉型が混在: {key}"

    def test_message_id_unique(self):
        """message_id は全メッセージで一意（公式 ver5.3 準拠の通し連番・F2/F3共通）

        旧形式 {id}_{round} は質問/回答で同一IDになり、ingest の doc_id 上書きで
        片方のメッセージが silently 消失していた（実データで 271→158 件に減少）。
        通し連番 {id}_{seq} は読み順で各メッセージに一意番号を振る。
        """
        from apps.backend.app.preliminary_review.knowledge.excel_reader import read_all_f2, read_all_f3
        for label, records in (("F3", read_all_f3()), ("F2", read_all_f2())):
            assert records, f"{label}レコードが読めない"
            message_ids = [r["message_id"] for r in records]
            assert len(message_ids) == len(set(message_ids)), (
                f"{label} message_id 重複: {len(message_ids)} 件中 一意 {len(set(message_ids))} 件"
            )

    def test_excel_reader_adds_sheet_name(self, tmp_path):
        """スキーマに sheet_name があれば各レコードに由来シートが付く（ver5.3・Step1-1）"""
        from openpyxl import Workbook
        from apps.backend.app.preliminary_review.knowledge.excel_reader import _read_excel_by_schema

        schema = {
            "sheet_name": "KNI_1G_01",
            "layout": {"data_start_row": 1},
            "loader_config": {"id_column": "A"},
            "fixed_columns": [{"key": "id", "col": "A", "dtype": "string"}],
            "output_model": {"flatten_qa": False},
            "meta_cells": {},
        }
        wb = Workbook()
        wb.active.append(["03_1G_01_0001"])
        xlsx = tmp_path / "mini_f3_sheetname.xlsx"
        wb.save(xlsx)

        records, _ = _read_excel_by_schema(schema, xlsx)

        assert len(records) == 1
        assert records[0]["sheet_name"] == "KNI_1G_01"

    def test_ver53_columns_cover_schemas(self):
        """全F2/F3スキーマの fixed_columns キーが ver5.3 正準列＋付帯列に収まっている（Step1）

        逆方向も検証：ver5.3 本体列のうち flatten_qa 生成分（message_id/message_content）
        を除くすべてが、各スキーマの fixed_columns に実在する。
        → schema YAML と VER53_SCHEMA の契約が乖離したらここで落ちる。
        """
        from apps.backend.app.preliminary_review.knowledge.excel_reader import (
            VER53_SCHEMA,
            _discover_schemas,
        )
        generated_keys = {"message_id", "message_content"}
        for ktype in ("f2", "f3"):
            columns, aux_columns = VER53_SCHEMA[ktype]
            allowed = set(columns) | set(aux_columns)
            required_fixed = set(columns) - generated_keys
            schemas = _discover_schemas(ktype)
            assert schemas, f"{ktype}スキーマが検出されない"
            for schema in schemas:
                keys = {c["key"] for c in schema.get("fixed_columns", [])}
                assert keys <= allowed, (
                    f"{ktype}/{schema['sheet_name']}: ver5.3契約外の列 {keys - allowed}"
                )
                assert required_fixed <= keys, (
                    f"{ktype}/{schema['sheet_name']}: ver5.3必須列の欠落 {required_fixed - keys}"
                )

    def test_to_ver53_rows_projection(self):
        """to_ver53_rows は正準列＋付帯列に射影し、欠損は既定値・余分キーは落とす（Step1）"""
        from apps.backend.app.preliminary_review.knowledge.excel_reader import (
            VER53_SCHEMA,
            to_ver53_rows,
        )
        # F3（デフォルト）
        f3_cols = set(VER53_SCHEMA["f3"][0]) | set(VER53_SCHEMA["f3"][1])
        records = [{
            "id": "03_1G_01_0001",
            "message_id": "03_1G_01_0001_01",
            "message_content": "確認します",
            "cost_category": "維持管理費",
            "sheet_name": "KNI_1G_01",
            "round": 1,
            "unexpected_key": "落とすべき",
        }]
        row = to_ver53_rows(records)[0]
        assert set(row.keys()) == f3_cols
        assert "unexpected_key" not in row
        assert row["cost_category"] == "維持管理費"
        assert row["submission_timing"] == ""   # 欠損は空文字
        assert row["round"] == 1                # round は数値のまま
        assert to_ver53_rows([{}])[0]["round"] == 0  # round の既定値は 0

        # F2（knowledge_type 指定）＝F2固有列に射影され、F3固有列は含まれない
        f2_cols = set(VER53_SCHEMA["f2"][0]) | set(VER53_SCHEMA["f2"][1])
        f2_row = to_ver53_rows([{
            "id": "02_1G_0001",
            "business_category": "工事情報",
            "cost_category": "落とすべき（F2にcost_categoryは無い）",
        }], knowledge_type="f2")[0]
        assert set(f2_row.keys()) == f2_cols
        assert f2_row["business_category"] == "工事情報"
        assert "cost_category" not in f2_row
        assert "reactor_type" not in f2_row  # F3固有の付帯列は混ざらない

    def test_supplement_returns_empty_for_denryoku(self):
        """電力ロールは補足資料を参照できない（Phase 3）"""
        from apps.backend.app.preliminary_review.knowledge.knowledge_loader import load_supplement
        result = load_supplement(caller_role="電力")
        assert result == []

    def test_supplement_returns_empty_when_datastore_not_configured(self):
        """データストアID未設定時は空リストにフォールバックする（Phase 3）"""
        import apps.backend.app.preliminary_review.knowledge.knowledge_loader as kl
        original = kl.VERTEX_SEARCH_SUPPLEMENT_DATASTORE_ID
        try:
            kl.VERTEX_SEARCH_SUPPLEMENT_DATASTORE_ID = ""
            result = kl.load_supplement(caller_role="NuRO")
            assert result == []
        finally:
            kl.VERTEX_SEARCH_SUPPLEMENT_DATASTORE_ID = original


# ─────────────────────────────────────────────────────────────────────────────
# 3. POST /api/review エンドポイント E2E テスト
# ─────────────────────────────────────────────────────────────────────────────

MOCK_MAPPINGS = [
    {"field_name": "計画実績区分", "cell_address": "C8",  "value": "実績",           "reasoning": ""},
    {"field_name": "電力会社",     "cell_address": "C5",  "value": "AA電力株式会社", "reasoning": ""},
    {"field_name": "総額",         "cell_address": "G18", "value": "1000",           "reasoning": ""},
    {"field_name": "総額",         "cell_address": "K18", "value": "1500",           "reasoning": ""},
]

MOCK_GEMINI_RESPONSE = json.dumps({
    "review_items": [
        {
            "field_name": "総額",
            "cell_address": "K18",
            "severity": "要確認",
            "comment": "計画値と実績値の乖離が50%です",
            "evidence": "計画差分: G18=1000, K18=1500",
            "knowledge_source": "計画差分",
        }
    ],
    "summary": "総額に大きな乖離があります。",
})


def _make_mock_firestore(mappings: list[dict] = MOCK_MAPPINGS, reviewed: bool = False):
    """Firestoreクライアントのモックを生成する"""
    doc = MagicMock()
    doc.exists = True
    doc.to_dict.return_value = {
        "session_id": "test-session-001",
        "utility_name": "AA電力",
        "frame_name": "frameB",
        "sheet_name": "MRC1",
        "mappings": mappings,
        "reviewed": reviewed,
    }
    db = MagicMock()
    db.collection.return_value.document.return_value.get.return_value = doc
    db.collection.return_value.document.return_value.collection.return_value.document.return_value.set = MagicMock()
    db.collection.return_value.document.return_value.update = MagicMock()
    return db


_REVIEW_FS_PATH   = "apps.backend.app.api.routes.review.get_firestore_client"
_UPLOAD_FS_PATH   = "apps.backend.app.api.routes.upload.get_firestore_client"
# ADK 移行後、call_gemini は adk/agents.py 内で使われるためモック先を変更
_GEMINI_PATH      = "apps.backend.app.preliminary_review.workflow.nodes.call_gemini"


@pytest.fixture
def client():
    """FastAPI テストクライアント（Firestore・Gemini をモック）"""
    # review.py が直接参照する名前をパッチする（呼び出し元モジュールへのパッチが必要）
    with patch(_REVIEW_FS_PATH, return_value=_make_mock_firestore()), \
         patch(_UPLOAD_FS_PATH, return_value=MagicMock()), \
         patch(_GEMINI_PATH, return_value=MOCK_GEMINI_RESPONSE):
        from apps.backend.app.api.main import app
        yield TestClient(app)


class TestReviewEndpoint:

    def test_review_success(self, client):
        """正常系: レビュー結果と mappings が返る"""
        res = client.post("/api/review", json={
            "session_id": "test-session-001",
            "utility_name": "AA電力",
            "sheet_name": "MRC1",
            "frame_name": "frameB",
        })
        assert res.status_code == 200
        data = res.json()
        assert "review_id" in data
        assert isinstance(data["review_items"], list)
        assert len(data["review_items"]) >= 1
        # mappings がレスポンスに含まれる（グリッド表示修正）
        assert "mappings" in data
        assert len(data["mappings"]) == len(MOCK_MAPPINGS)

    def test_review_session_not_found(self):
        """セッション不存在は 404 を返す"""
        doc = MagicMock()
        doc.exists = False
        db = MagicMock()
        db.collection.return_value.document.return_value.get.return_value = doc

        with patch(_REVIEW_FS_PATH, return_value=db), \
             patch(_GEMINI_PATH, return_value=MOCK_GEMINI_RESPONSE):
            from apps.backend.app.api.main import app
            c = TestClient(app)
            res = c.post("/api/review", json={
                "session_id": "no-such-session",
                "utility_name": "AA電力",
            })
        assert res.status_code == 404

    def test_review_empty_mappings(self):
        """mappings が空のセッションは 400 を返す"""
        db = _make_mock_firestore(mappings=[])
        with patch(_REVIEW_FS_PATH, return_value=db), \
             patch(_GEMINI_PATH, return_value=MOCK_GEMINI_RESPONSE):
            from apps.backend.app.api.main import app
            c = TestClient(app)
            res = c.post("/api/review", json={
                "session_id": "empty-session",
                "utility_name": "AA電力",
            })
        assert res.status_code == 400

    def test_review_response_includes_severity(self, client):
        """指摘の severity が ReviewItem に含まれる"""
        res = client.post("/api/review", json={
            "session_id": "test-session-001",
            "utility_name": "AA電力",
        })
        assert res.status_code == 200
        items = res.json()["review_items"]
        for item in items:
            assert item["severity"] in ("要確認", "AIからの指摘")


# ─────────────────────────────────────────────────────────────────────────────
# 4. POST /api/review/{id}/feedback エンドポイント テスト
# ─────────────────────────────────────────────────────────────────────────────

def _make_fs_with_review():
    """フィードバック用Firestoreモック（毎回フレッシュなmockを返す）"""
    review_doc = MagicMock()
    review_doc.reference.update = MagicMock()
    db = MagicMock()
    db.collection_group.return_value.where.return_value.limit.return_value.stream.return_value = iter([review_doc])
    return db, review_doc


class TestFeedbackEndpoint:

    def test_accept_returns_saved(self):
        """承諾（accept）は "saved" を返し Firestore に書き込む"""
        db, review_doc = _make_fs_with_review()
        with patch(_REVIEW_FS_PATH, return_value=db):
            from apps.backend.app.api.main import app
            c = TestClient(app)
            res = c.post("/api/review/rev-001/feedback", json={
                "item_id": "review_001",
                "decision": "accept",
                "comment": "",
            })
        assert res.status_code == 200
        assert res.json()["status"] == "saved"
        review_doc.reference.update.assert_called_once()

    def test_reject_returns_discarded_and_increments_decided_count(self):
        """棄却（reject）は "discarded" を返し decided_count をインクリメントする"""
        review_doc = MagicMock()
        review_doc.reference.update = MagicMock()
        db = MagicMock()
        db.collection_group.return_value.where.return_value.limit.return_value.stream.return_value = iter([review_doc])
        with patch(_REVIEW_FS_PATH, return_value=db):
            from apps.backend.app.api.main import app
            c = TestClient(app)
            res = c.post("/api/review/rev-001/feedback", json={
                "item_id": "review_001",
                "decision": "reject",
                "comment": "不要と判断",
            })
        assert res.status_code == 200
        assert res.json()["status"] == "discarded"
        # 棄却は decided_count のみインクリメント（指摘内容は保存しない）
        review_doc.reference.update.assert_called_once()

    def test_invalid_decision_returns_400(self):
        """無効な decision 値は 400 を返す"""
        db = MagicMock()
        with patch(_REVIEW_FS_PATH, return_value=db):
            from apps.backend.app.api.main import app
            c = TestClient(app)
            res = c.post("/api/review/rev-001/feedback", json={
                "item_id": "review_001",
                "decision": "maybe",
            })
        assert res.status_code == 400


# ─────────────────────────────────────────────────────────────────────────────
# 5. GET /api/review/sessions エンドポイント テスト
# ─────────────────────────────────────────────────────────────────────────────

class TestSessionsEndpoint:

    def test_returns_unreviewed_sessions(self):
        """未レビューセッションの一覧が返る"""
        from datetime import datetime, timezone

        mock_doc = MagicMock()
        mock_doc.id = "sess-001"
        mock_doc.to_dict.return_value = {
            "session_id": "sess-001",
            "utility_name": "AA電力",
            "frame_name": "frameB",
            "sheet_name": "MRC1",
            "created_at": datetime(2026, 5, 15, tzinfo=timezone.utc),
            "reviewed": False,
        }

        db = MagicMock()
        db.collection.return_value.where.return_value.order_by.return_value.limit.return_value.stream.return_value = iter([mock_doc])

        with patch(_REVIEW_FS_PATH, return_value=db):
            from apps.backend.app.api.main import app
            c = TestClient(app)
            res = c.get("/api/review/sessions")

        assert res.status_code == 200
        sessions = res.json()
        assert len(sessions) == 1
        assert sessions[0]["session_id"] == "sess-001"
        assert sessions[0]["utility_name"] == "AA電力"
        assert sessions[0]["reviewed"] is False

    def test_returns_empty_when_no_sessions(self):
        """セッションがない場合は空リストを返す"""
        db = MagicMock()
        db.collection.return_value.where.return_value.order_by.return_value.limit.return_value.stream.return_value = iter([])

        with patch(_REVIEW_FS_PATH, return_value=db):
            from apps.backend.app.api.main import app
            c = TestClient(app)
            res = c.get("/api/review/sessions")

        assert res.status_code == 200
        assert res.json() == []


# ─────────────────────────────────────────────────────────────────────────────
# 6. review_workbook — ワークブック一括レビューの統括関数（Step2）
# ─────────────────────────────────────────────────────────────────────────────

class TestReviewWorkbook:
    """review_workbook() のオーケストレーションを検証する。

    Excel復元・クエリ文脈導出・run_review はすべてモックし、
    「シート列挙 → 復元 → スキップ判定 → run_review 呼び出し → 結果集約」
    の制御フローだけをGCP接続なしで検証する（特定データファイルに依存しない）。
    """

    _CTX = {"fee_type": "解体撤去費", "reactor_type": "PWR", "utility_name": "AA電力"}

    def _run(self, tmp_path, *, sheets=("MRC1", "MRC2"), mappings_by_sheet=None,
             utility_name=None, sheet_names=None, ctx=None):
        """モックを組んで review_workbook を実行し (結果, run_reviewモック) を返す。"""
        import asyncio
        from apps.backend.app.preliminary_review import agent as ra

        excel = tmp_path / "result.xlsx"
        excel.write_bytes(b"dummy")  # 存在チェック用（読み込みはモックする）

        default_mappings = [{"field_name": "工事件名", "cell_address": "C5",
                             "value": "解体工事", "reasoning": ""}]
        mappings_by_sheet = mappings_by_sheet or {s: default_mappings for s in sheets}

        def _fake_reconstruct(path, frame, sheet):
            result = mappings_by_sheet.get(sheet, [])
            if isinstance(result, Exception):
                raise result
            return result

        mock_run_review = AsyncMock(return_value=([], [{"tool": "Tool1", "count": 0}]))
        with patch.object(ra, "list_frame_sheets", return_value=list(sheets)), \
             patch.object(ra, "reconstruct_mappings_from_excel", side_effect=_fake_reconstruct), \
             patch.object(ra, "derive_query_context", return_value=dict(ctx or self._CTX)), \
             patch.object(ra, "run_review", mock_run_review):
            result = asyncio.run(ra.review_workbook(
                excel_path=excel,
                frame_name="frameB",
                sheet_names=sheet_names,
                utility_name=utility_name,
            ))
        return result, mock_run_review

    def test_reviews_all_frame_sheets(self, tmp_path):
        """シート未指定なら frame の全シートを一括レビューする"""
        result, mock_rr = self._run(tmp_path)
        assert set(result["sheets"].keys()) == {"MRC1", "MRC2"}
        assert result["skipped_sheets"] == []
        assert mock_rr.await_count == 2

    def test_passes_query_context_and_utility(self, tmp_path):
        """基本情報シート由来の費目・炉型・会社が run_review に渡る"""
        result, mock_rr = self._run(tmp_path)
        assert result["utility_name"] == "AA電力"  # ctx から自動解決
        for call in mock_rr.await_args_list:
            assert call.kwargs["reactor_type"] == "PWR"
            assert call.kwargs["fee_type"] == "解体撤去費"
            assert call.kwargs["utility_name"] == "AA電力"

    def test_explicit_utility_overrides_context(self, tmp_path):
        """utility_name を明示指定した場合は Excel 由来より優先する"""
        result, mock_rr = self._run(tmp_path, utility_name="BB電力")
        assert result["utility_name"] == "BB電力"
        assert mock_rr.await_args_list[0].kwargs["utility_name"] == "BB電力"

    def test_utility_fallback_when_context_empty(self, tmp_path):
        """会社名が引数にも Excel にも無ければ「不明電力」で続行する"""
        ctx = {"fee_type": None, "reactor_type": None, "utility_name": None}
        result, _ = self._run(tmp_path, ctx=ctx)
        assert result["utility_name"] == "不明電力"

    def test_skips_empty_and_failing_sheets(self, tmp_path):
        """mappings が空・復元例外のシートはスキップし、残りはレビューする"""
        mappings = {
            "MRC1": [{"field_name": "工事件名", "cell_address": "C5",
                      "value": "解体工事", "reasoning": ""}],
            "MRC2": [],                        # 空 → スキップ
            "MRC3": FileNotFoundError("様式定義なし"),  # 例外 → スキップ
        }
        result, mock_rr = self._run(tmp_path, sheets=("MRC1", "MRC2", "MRC3"),
                                    mappings_by_sheet=mappings)
        assert set(result["sheets"].keys()) == {"MRC1"}
        assert sorted(result["skipped_sheets"]) == ["MRC2", "MRC3"]
        assert mock_rr.await_count == 1

    def test_explicit_sheet_names_restrict_targets(self, tmp_path):
        """sheet_names を指定した場合はそのシートだけレビューする"""
        result, mock_rr = self._run(tmp_path, sheet_names=["MRC2"])
        assert set(result["sheets"].keys()) == {"MRC2"}
        assert mock_rr.await_count == 1

    def test_missing_excel_raises(self, tmp_path):
        """存在しないExcelパスは FileNotFoundError"""
        import asyncio
        from apps.backend.app.preliminary_review import agent as ra
        with pytest.raises(FileNotFoundError):
            asyncio.run(ra.review_workbook(excel_path=tmp_path / "nai.xlsx"))

    def test_no_frame_config_raises(self, tmp_path):
        """frame のシート定義が config に無ければ ValueError"""
        import asyncio
        from apps.backend.app.preliminary_review import agent as ra
        excel = tmp_path / "result.xlsx"
        excel.write_bytes(b"dummy")
        with patch.object(ra, "list_frame_sheets", return_value=[]):
            with pytest.raises(ValueError):
                asyncio.run(ra.review_workbook(excel_path=excel, frame_name="nashi"))


# ─────────────────────────────────────────────────────────────────────────────
# 7. Excel読み取りの行独立性 — ffill捏造の再発防止（2026-07-04・§1-16）
# ─────────────────────────────────────────────────────────────────────────────

class TestRowIsolationInExcelReader:
    """_read_excel_by_schema の「空セルは空のまま」を固定するテスト。

    旧実装は全列を無条件 ffill（下方向前埋め）しており、上の行のメッセージが
    別IDのスレッドに複製される（F3 209→271件・F2 56→86件に水増し）バグがあった。
    結合セルは _expand_merged_cells() が結合範囲に限定して展開する。
    """

    _SCHEMA = {
        "layout": {"data_start_row": 2},
        "loader_config": {"id_column": "A"},
        "fixed_columns": [
            {"key": "id",       "col": "A", "dtype": "string"},
            {"key": "optional", "col": "B", "dtype": "string"},
        ],
        "repeating_qa_columns": {
            "start_col": "C", "col_per_round": 2, "max_rounds": 2,
            "fields": [
                {"key": "nuro_comment",   "col_offset": 0},
                {"key": "denryoku_reply", "col_offset": 1},
            ],
        },
        "output_model": {"flatten_qa": True},
        "meta_cells": {},
    }

    def _write_xlsx(self, tmp_path, rows, merges=()):
        """テスト用の小さなExcelを作る（1行目=ヘッダー・2行目以降=データ）。"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "任意列", "確認1", "回答1", "確認2", "回答2"])
        for r in rows:
            ws.append(r)
        for m in merges:
            ws.merge_cells(m)
        path = tmp_path / "mini_knowledge.xlsx"
        wb.save(path)
        return path

    def test_no_message_fabrication_from_previous_row(self, tmp_path):
        """上の行に2往復あっても、1往復しかない次の行にメッセージが複製されない"""
        from apps.backend.app.preliminary_review.knowledge.excel_reader import _read_excel_by_schema
        path = self._write_xlsx(tmp_path, [
            ["ID-001", "資料URL", "確認A", "回答A", "確認B", "回答B"],  # 2往復
            ["ID-002", "",        "確認C", "回答C", "",      ""],       # 1往復のみ
        ])
        records, _ = _read_excel_by_schema(self._SCHEMA, path)

        by_id = {}
        for r in records:
            by_id.setdefault(r["id"], []).append(r)
        assert len(by_id["ID-001"]) == 4
        assert len(by_id["ID-002"]) == 2, (
            f"ID-002 は2メッセージのはず（ffill捏造の再発）: "
            f"{[r['message_content'] for r in by_id['ID-002']]}"
        )
        assert [r["message_content"] for r in by_id["ID-002"]] == ["確認C", "回答C"]

    def test_no_fixed_column_bleed_from_previous_row(self, tmp_path):
        """意図的に空の固定列に、上の行の値（URL等）が染み出さない"""
        from apps.backend.app.preliminary_review.knowledge.excel_reader import _read_excel_by_schema
        path = self._write_xlsx(tmp_path, [
            ["ID-001", "https://example.com/a.xlsx", "確認A", "回答A", "", ""],
            ["ID-002", "",                           "確認B", "回答B", "", ""],
        ])
        records, _ = _read_excel_by_schema(self._SCHEMA, path)
        id2 = [r for r in records if r["id"] == "ID-002"]
        assert all(r["optional"] == "" for r in id2), (
            f"ID-002 の optional は空のはず: {id2[0]['optional']!r}"
        )

    def test_merged_cells_are_expanded(self, tmp_path):
        """結合セルは結合範囲に限定してアンカー値が展開される（fill_downの本来の意図）"""
        from apps.backend.app.preliminary_review.knowledge.excel_reader import _read_excel_by_schema
        path = self._write_xlsx(tmp_path, [
            ["ID-001", "共通資料", "確認A", "回答A", "", ""],
            ["ID-002", "",         "確認B", "回答B", "", ""],
            ["ID-003", "",         "確認C", "回答C", "", ""],
        ], merges=["B2:B3"])  # ID-001とID-002の任意列を結合（ID-003は結合外）
        records, _ = _read_excel_by_schema(self._SCHEMA, path)
        by_id = {r["id"]: r for r in records}
        assert by_id["ID-001"]["optional"] == "共通資料"
        assert by_id["ID-002"]["optional"] == "共通資料"  # 結合範囲内＝展開される
        assert by_id["ID-003"]["optional"] == ""          # 結合範囲外＝空のまま


# ─────────────────────────────────────────────────────────────────────────────
# 8. Reranking（Ranking API）— surfacing底上げ＋ガードのスコア統合（Step5・§1-19）
# ─────────────────────────────────────────────────────────────────────────────

class TestRerank:
    """knowledge_loader._rerank のオーケストレーションを検証する（GCP接続なし）。

    RankServiceClient をモックし、並べ替え・スコア付与・安全フォールバックの
    制御フローだけを確認する。
    """

    def _fake_rank_response(self, ordered):
        """ordered=[(id, score),...] から RankResponse 相当のモックを返す。"""
        resp = MagicMock()
        resp.records = [MagicMock(id=str(i), score=s) for i, s in ordered]
        return resp

    def _records(self):
        return [
            {"_doc_id": "A", "message_content": "解体撤去の費用内訳"},
            {"_doc_id": "B", "message_content": "放射性廃棄物の処分"},
            {"_doc_id": "C", "message_content": "解体工法の選定基準"},
        ]

    def test_reorders_by_score_and_attaches_score(self):
        """スコア降順に並べ替わり、各レコードに _rerank_score が付く"""
        from apps.backend.app.preliminary_review.knowledge import knowledge_loader as kl
        client = MagicMock()
        # 入力順 A,B,C を C,A,B に並べ替え（idは入力インデックス0,1,2）
        client.rank.return_value = self._fake_rank_response([(2, 0.91), (0, 0.55), (1, 0.10)])
        client.ranking_config_path.return_value = "rc"
        with patch.object(kl, "RERANK_ENABLED", True), \
             patch.object(kl, "GCP_PROJECT_ID", "proj"), \
             patch.object(kl, "_get_rank_client", return_value=client):
            out = kl._rerank("解体撤去費", self._records())
        assert [r["_doc_id"] for r in out] == ["C", "A", "B"]
        assert out[0]["_rerank_score"] == 0.91
        assert all("_rerank_score" in r for r in out)

    def test_api_error_falls_back_to_original_order(self):
        """API エラー時は元の順序をそのまま返す（検索を止めない）"""
        from apps.backend.app.preliminary_review.knowledge import knowledge_loader as kl
        from google.api_core.exceptions import GoogleAPICallError
        client = MagicMock()
        client.rank.side_effect = GoogleAPICallError("boom")
        client.ranking_config_path.return_value = "rc"
        with patch.object(kl, "RERANK_ENABLED", True), \
             patch.object(kl, "GCP_PROJECT_ID", "proj"), \
             patch.object(kl, "_get_rank_client", return_value=client):
            out = kl._rerank("q", self._records())
        assert [r["_doc_id"] for r in out] == ["A", "B", "C"]
        assert all("_rerank_score" not in r for r in out)

    def test_disabled_skips_rank_call(self):
        """RERANK_ENABLED=false なら rank() を呼ばず素通し"""
        from apps.backend.app.preliminary_review.knowledge import knowledge_loader as kl
        client = MagicMock()
        with patch.object(kl, "RERANK_ENABLED", False), \
             patch.object(kl, "GCP_PROJECT_ID", "proj"), \
             patch.object(kl, "_get_rank_client", return_value=client):
            out = kl._rerank("q", self._records())
        client.rank.assert_not_called()
        assert [r["_doc_id"] for r in out] == ["A", "B", "C"]

    def test_missing_response_records_are_appended(self):
        """rank応答が一部欠けても取りこぼさず全件返す（安全網）"""
        from apps.backend.app.preliminary_review.knowledge import knowledge_loader as kl
        client = MagicMock()
        client.rank.return_value = self._fake_rank_response([(2, 0.9)])  # C のみ返る
        client.ranking_config_path.return_value = "rc"
        with patch.object(kl, "RERANK_ENABLED", True), \
             patch.object(kl, "GCP_PROJECT_ID", "proj"), \
             patch.object(kl, "_get_rank_client", return_value=client):
            out = kl._rerank("q", self._records())
        assert {r["_doc_id"] for r in out} == {"A", "B", "C"}
        assert out[0]["_doc_id"] == "C"

    def test_duplicate_ids_do_not_double_count_or_drop(self):
        """重複idは1回だけ採用し、欠けたレコードは末尾に温存する（二重採用・取りこぼし防止）"""
        from apps.backend.app.preliminary_review.knowledge import knowledge_loader as kl
        client = MagicMock()
        # id "0" が2回・"1" が欠落。素朴な長さ一致判定だと B を落とし A を二重採用してしまう
        client.rank.return_value = self._fake_rank_response([(0, 0.9), (0, 0.8), (2, 0.5)])
        client.ranking_config_path.return_value = "rc"
        with patch.object(kl, "RERANK_ENABLED", True), \
             patch.object(kl, "GCP_PROJECT_ID", "proj"), \
             patch.object(kl, "_get_rank_client", return_value=client):
            out = kl._rerank("q", self._records())
        assert [r["_doc_id"] for r in out] == ["A", "C", "B"]  # A1回・C・欠落Bを末尾温存
        assert len(out) == 3

    def test_negative_and_out_of_range_ids_are_skipped(self):
        """負数・範囲外idは無視し、欠けたレコードは末尾に温存する"""
        from apps.backend.app.preliminary_review.knowledge import knowledge_loader as kl
        client = MagicMock()
        client.rank.return_value = self._fake_rank_response([(-1, 0.9), (99, 0.8), (1, 0.5)])
        client.ranking_config_path.return_value = "rc"
        with patch.object(kl, "RERANK_ENABLED", True), \
             patch.object(kl, "GCP_PROJECT_ID", "proj"), \
             patch.object(kl, "_get_rank_client", return_value=client):
            out = kl._rerank("q", self._records())
        assert {r["_doc_id"] for r in out} == {"A", "B", "C"}
        assert out[0]["_doc_id"] == "B"  # 有効idの1のみ先頭、A/Cは末尾温存

    def test_caption_used_as_content_for_supplement(self):
        """message_content が無い補足資料（Tool4）は caption を検索対象テキストに使う"""
        from apps.backend.app.preliminary_review.knowledge import knowledge_loader as kl
        captured = {}

        def _capture_rank(request):
            captured["contents"] = [r.content for r in request.records]
            return self._fake_rank_response([(0, 0.9)])

        client = MagicMock()
        client.rank.side_effect = _capture_rank
        client.ranking_config_path.return_value = "rc"
        recs = [{"_doc_id": "S", "caption": "格納容器の外観写真。腐食が確認できる。"}]
        with patch.object(kl, "RERANK_ENABLED", True), \
             patch.object(kl, "GCP_PROJECT_ID", "proj"), \
             patch.object(kl, "_get_rank_client", return_value=client):
            kl._rerank("q", recs)
        assert captured["contents"] == ["格納容器の外観写真。腐食が確認できる。"]

    def test_search_invokes_rerank(self):
        """_search が検索結果を _rerank に通す配線を固定（呼び出し削除の回帰検知）"""
        from apps.backend.app.preliminary_review.knowledge import knowledge_loader as kl
        search_client = MagicMock()
        search_client.search.return_value = MagicMock(results=["r1", "r2"])
        sentinel = [{"_doc_id": "X"}]
        with patch.object(kl, "GCP_PROJECT_ID", "proj"), \
             patch.object(kl, "_get_search_client", return_value=search_client), \
             patch.object(kl, "_to_record", side_effect=lambda r: {"_doc_id": r}), \
             patch.object(kl, "_rerank", return_value=sentinel) as mock_rerank:
            out = kl._search("ds", "解体撤去費")
        mock_rerank.assert_called_once()
        # _search は _rerank の戻り値をそのまま返す
        assert out is sentinel


class TestRelevanceGuardWithScore:
    """_record_relevant の Reranking スコア統合（§1-18 の偽陽性排除）。"""

    def test_f2_score_above_threshold_is_relevant(self):
        """高スコアなら費目と語を共有しなくても関連＝スコア経路を証明（トークン経路では拾えない）"""
        from apps.backend.app.preliminary_review import review_logic as rl
        # 内容は費目「施設解体一解体費」と2文字トークンを一切共有しない（＝旧トークン判定なら False）
        rec = {"message_content": "基礎撤去に伴う産業廃棄物の処理手順", "_rerank_score": 0.72}
        with patch.object(rl, "RERANK_GUARD_F2_THRESHOLD", 0.35):
            assert rl._record_relevant("施設解体一解体費", rec) is True

    def test_f2_low_score_rejected_even_if_tokens_collide(self):
        """『放射線管理費』×『放射性廃棄物』は2文字"放射"で一致するが、低スコアで不採用"""
        from apps.backend.app.preliminary_review import review_logic as rl
        rec = {"message_content": "一次系配管解体に伴う放射性廃棄物の追加発生", "_rerank_score": 0.08}
        with patch.object(rl, "RERANK_GUARD_F2_THRESHOLD", 0.35):
            # スコアが閾値未満なので False（字面トークンなら True になってしまうケース）
            assert rl._record_relevant("放射線管理費", rec) is False

    def test_f2_without_score_falls_back_to_content_tokens(self):
        """スコアが無ければ従来の内容語トークン判定にフォールバック"""
        from apps.backend.app.preliminary_review import review_logic as rl
        rec = {"message_content": "解体工法の選定基準"}  # _rerank_score なし
        assert rl._record_relevant("施設解体一解体費", rec) is True

    def test_f3_with_fee_type_ignores_score(self):
        """費目を持つF3はスコアに関係なく従来どおり費目語で判定"""
        from apps.backend.app.preliminary_review import review_logic as rl
        rec = {"cost_category": "解体撤去費", "fee_type": "解体撤去費", "_rerank_score": 0.01}
        assert rl._record_relevant("施設解体一解体費", rec) is True


class TestSettingsEnvParsing:
    """settings の環境変数パーサ（不正値でアプリ起動を止めない・一般的な真値を許容）。"""

    def test_env_bool_accepts_common_truthy(self, monkeypatch):
        from apps.backend.app.core.settings import _env_bool
        for v in ("1", "true", "True", "YES", "on"):
            monkeypatch.setenv("X_RERANK_T", v)
            assert _env_bool("X_RERANK_T", False) is True
        for v in ("0", "false", "no", "off", "maybe"):
            monkeypatch.setenv("X_RERANK_T", v)
            assert _env_bool("X_RERANK_T", True) is False

    def test_env_bool_default_when_unset(self, monkeypatch):
        from apps.backend.app.core.settings import _env_bool
        monkeypatch.delenv("X_RERANK_UNSET", raising=False)
        assert _env_bool("X_RERANK_UNSET", True) is True

    def test_env_float_bad_value_falls_back(self, monkeypatch):
        """ロケールカンマ等の不正値でも例外を投げず既定値に退避（import時クラッシュ防止）"""
        from apps.backend.app.core.settings import _env_float
        monkeypatch.setenv("X_RERANK_TH", "0,15")
        assert _env_float("X_RERANK_TH", 0.15) == 0.15
        monkeypatch.setenv("X_RERANK_TH", "0.42")
        assert _env_float("X_RERANK_TH", 0.15) == 0.42


class TestReanchorReviewItems:
    """空欄項目への指摘の番地補正（様式定義駆動・#番地誤爆の是正）。"""

    _MAPS = [{"field_name": "計画実績区分", "cell_address": "C8", "value": "計画", "reasoning": ""}]

    def _item(self, field, cell):
        from apps.backend.app.api.models import ReviewItem
        return ReviewItem(item_id="", field_name=field, cell_address=cell,
                          severity="要確認", comment="", evidence="", knowledge_source="AI知見")

    def test_empty_field_misanchor_is_corrected(self):
        """空欄項目(実施費用低減策)の指摘が可視セルC6に誤爆 → 様式定義でG24に補正"""
        from apps.backend.app.preliminary_review.review_logic import reanchor_review_items
        items = [self._item("実施費用低減策", "C6")]
        reanchor_review_items(items, "frameB", "MRC1", self._MAPS)
        assert items[0].cell_address == "G24"

    def test_plan_actual_prefers_actual_for_jisseki(self):
        """実績提出なら plan_actual 項目は actual 列(K24)へ補正"""
        from apps.backend.app.preliminary_review.review_logic import reanchor_review_items
        maps = [{"field_name": "計画実績区分", "cell_address": "C8", "value": "実績"}]
        items = [self._item("実施費用低減策", "C6")]
        reanchor_review_items(items, "frameB", "MRC1", maps)
        assert items[0].cell_address == "K24"

    def test_valid_cell_is_left_untouched(self):
        """既に定義セル(G24)に付いている指摘は温存（不要な補正をしない）"""
        from apps.backend.app.preliminary_review.review_logic import reanchor_review_items
        items = [self._item("実施費用低減策", "G24")]
        reanchor_review_items(items, "frameB", "MRC1", self._MAPS)
        assert items[0].cell_address == "G24"

    def test_field_in_two_sections_keeps_valid_cell(self):
        """複数セクションに定義される炉型(C7とG9)は、どちらの有効セルも補正しない"""
        from apps.backend.app.preliminary_review.review_logic import reanchor_review_items
        for cell in ("C7", "G9"):
            items = [self._item("炉型", cell)]
            reanchor_review_items(items, "frameB", "MRC1", self._MAPS)
            assert items[0].cell_address == cell, f"{cell} は温存されるべき"

    def test_tabular_field_is_left_untouched(self):
        """様式定義一覧に無い表フィールドは温存（元番地が正しいため触らない）"""
        from apps.backend.app.preliminary_review.review_logic import reanchor_review_items
        items = [self._item("解体機器表_30_計画_費用", "J30")]
        reanchor_review_items(items, "frameB", "MRC1", self._MAPS)
        assert items[0].cell_address == "J30"

    def test_unknown_field_is_left_untouched(self):
        """様式定義に無い field名（LLMの言い換え等）は温存"""
        from apps.backend.app.preliminary_review.review_logic import reanchor_review_items
        items = [self._item("費用低減の記載について", "C6")]
        reanchor_review_items(items, "frameB", "MRC1", self._MAPS)
        assert items[0].cell_address == "C6"


class TestHumanizeEvidenceRefs:
    """参照番号（[F3own#N] 等）→ 実出典表記（シート・メッセージID）への置換。

    実レコードは message_id 列を持たず、一意なメッセージIDは _doc_id
    （通し連番 {id}_{seq:02d}）である。基底 id はスレッド共通のため、_doc_id を正とする。
    """

    _F3_OWN = [
        {"sheet_name": "KNI_1G_01", "_doc_id": "03_KT_1G_01_0003_02", "id": "03_KT_1G_01_0003", "utility_name": "関東電力"},
        {"sheet_name": "KNI_1G_02", "_doc_id": "03_KT_1G_02_0005_02", "id": "03_KT_1G_02_0005", "utility_name": "関東電力"},
        # 上と同スレッド（基底 id 共通）だが別メッセージ（seq 03）。潰してはいけない。
        {"sheet_name": "KNI_1G_01", "_doc_id": "03_KT_1G_01_0003_03", "id": "03_KT_1G_01_0003", "utility_name": "関東電力"},
    ]
    _F3_ALL = [
        {"sheet_name": "HKD_1G", "_doc_id": "03_HK_1G_0002_02", "id": "03_HK_1G_0002", "utility_name": "北の海電力"},
    ]
    _F2 = [
        {"sheet_name": "KNS_1G", "_doc_id": "03_KS_1G_0001_02", "id": "03_KS_1G_0001"},
    ]

    def _item(self, evidence, comment="", source="F3"):
        from apps.backend.app.api.models import ReviewItem
        return ReviewItem(item_id="", field_name="f", cell_address="C6",
                          severity="要確認", comment=comment, evidence=evidence,
                          knowledge_source=source)

    def _run(self, items):
        from apps.backend.app.preliminary_review.review_logic import humanize_evidence_refs
        return humanize_evidence_refs(items, self._F2, self._F3_OWN, self._F3_ALL)

    def test_f3own_ref_replaced_with_source(self):
        """[F3own#1] が シート＋メッセージID(_doc_id)＋会社名 の出典表記になる"""
        items = self._run([self._item("過去事例 [F3own#1] を参照")])
        assert items[0].evidence == "過去事例 【F3ナレッジ（自社：関東電力）｜シートKNI_1G_01｜メッセージID 03_KT_1G_01_0003_02】 を参照"

    def test_f3all_and_f2_labels(self):
        """F3all は他社表記・F2 は NuRO内知見表記になる"""
        items = self._run([self._item("[F3all#1] と [F2#1]")])
        assert "【F3ナレッジ（他社：北の海電力）｜シートHKD_1G｜メッセージID 03_HK_1G_0002_02】" in items[0].evidence
        assert "【F2ナレッジ（NuRO内知見）｜シートKNS_1G｜メッセージID 03_KS_1G_0001_02】" in items[0].evidence
        assert "#" not in items[0].evidence

    def test_out_of_range_ref_kept_verbatim(self):
        """辿れない参照番号（範囲外）は原文温存（安全側）"""
        items = self._run([self._item("[F3own#9] 参照")])
        assert items[0].evidence == "[F3own#9] 参照"

    def test_duplicate_same_ref_collapsed(self):
        """同一事例を複数番号で引いた連なりは同一表記を1件に畳む"""
        items = self._run([self._item("根拠あり ([F3own#1], [F3own#1], [F3own#1])")])
        assert items[0].evidence == "根拠あり (【F3ナレッジ（自社：関東電力）｜シートKNI_1G_01｜メッセージID 03_KT_1G_01_0003_02】)"

    def test_same_thread_distinct_messages_not_collapsed(self):
        """基底 id が同じでも別メッセージ（_doc_id 相違）は畳まず両方残す"""
        items = self._run([self._item("([F3own#1]、[F3own#3])")])
        assert "03_KT_1G_01_0003_02" in items[0].evidence
        assert "03_KT_1G_01_0003_03" in items[0].evidence
        # 2件が別表記として残る（区切りで連結）
        assert items[0].evidence.count("【") == 2

    def test_mixed_run_dedup_preserves_order(self):
        """連なり内で重複だけ畳み、異なる出典は元順序を保つ"""
        items = self._run([self._item("[F3own#1], [F3own#2], [F3own#1]")])
        assert items[0].evidence == (
            "【F3ナレッジ（自社：関東電力）｜シートKNI_1G_01｜メッセージID 03_KT_1G_01_0003_02】, "
            "【F3ナレッジ（自社：関東電力）｜シートKNI_1G_02｜メッセージID 03_KT_1G_02_0005_02】"
        )

    def test_record_without_locator_kept_verbatim(self):
        """由来情報（シート・ID）を全く持たないレコードへの参照は原文温存"""
        from apps.backend.app.preliminary_review.review_logic import humanize_evidence_refs
        items = [self._item("[F3own#1] 参照")]
        humanize_evidence_refs(items, [], [{"utility_name": "関東電力"}], [])
        assert items[0].evidence == "[F3own#1] 参照"

    def test_comment_refs_also_replaced(self):
        """comment 内の参照番号も置換される（プロンプトは comment への引用も許可）"""
        items = self._run([self._item("根拠なし", comment="同種案件（[F3all#1]）で確認済み")])
        assert "【F3ナレッジ（他社：北の海電力）" in items[0].comment

    def test_ref_without_brackets_replaced(self):
        """角括弧なしの表記ゆれ（F3own#2）も置換される"""
        items = self._run([self._item("F3own#2 を参照")])
        assert items[0].evidence == "【F3ナレッジ（自社：関東電力）｜シートKNI_1G_02｜メッセージID 03_KT_1G_02_0005_02】 を参照"

    def test_no_refs_evidence_unchanged(self):
        """参照番号を含まない evidence（AI判断等）は不変"""
        text = "AI判断（ナレッジ参照なし）：記載が不足"
        items = self._run([self._item(text, source="AI知見")])
        assert items[0].evidence == text

    def test_returns_same_list_inplace(self):
        """guard/reanchor と同じ規約：同一リストを in-place 更新して返す"""
        src = [self._item("[F3own#1]")]
        result = self._run(src)
        assert result is src


class TestGenerateMissingEntryItems:
    """記載必須欄の空欄検出（決定論ルール・criteria YAML の opt-in 宣言駆動）。"""

    _REQ_FIELDS = ["工事件名", "実施費用低減策", "説明"]
    _REQ_TABLE = {"解体機器表": {"計画": ["計画_費用"], "実績": ["実績_費用"]}}

    def _map(self, field, cell, value):
        return {"field_name": field, "cell_address": cell, "value": value, "reasoning": ""}

    def _run(self, mappings, fields=None, table=None):
        from apps.backend.app.preliminary_review.review_logic import _generate_missing_entry_items
        return _generate_missing_entry_items(
            mappings, "frameB", "MRC1",
            self._REQ_FIELDS if fields is None else fields,
            self._REQ_TABLE if table is None else table,
        )

    def test_absent_required_field_flagged_at_config_cell(self):
        """mappings に載っていない必須項目（未転記）は様式定義のセルで指摘される"""
        items = self._run([self._map("計画実績区分", "C8", "計画")], table={})
        by_field = {i.field_name: i for i in items}
        assert by_field["実施費用低減策"].cell_address == "G24"
        assert by_field["説明"].cell_address == "G25"
        assert by_field["実施費用低減策"].severity == "要確認"
        assert "記入してください" in by_field["実施費用低減策"].comment

    def test_empty_value_flagged_and_filled_not_flagged(self):
        """空値で載っている項目は指摘・記載済みの項目は指摘しない"""
        maps = [
            self._map("計画実績区分", "C8", "計画"),
            self._map("工事件名", "C6", "館山2号機 解体撤去工事"),  # 記載あり
            self._map("実施費用低減策", "G24", "  "),               # 空白のみ＝空欄
        ]
        items = self._run(maps, table={})
        fields = {i.field_name for i in items}
        assert "工事件名" not in fields
        assert "実施費用低減策" in fields

    def test_jisseki_prefers_actual_cell(self):
        """実績提出なら plan_actual 項目は実績セル(K24)で指摘される"""
        items = self._run([self._map("計画実績区分", "C8", "実績")], table={})
        by_field = {i.field_name: i.cell_address for i in items}
        assert by_field["実施費用低減策"] == "K24"

    def test_unknown_declaration_skipped(self):
        """様式定義に無い必須宣言（設定ミス）は黙ってスキップ"""
        items = self._run([], fields=["存在しない項目"], table={})
        assert items == []

    def test_table_empty_cells_aggregated_per_column(self):
        """表のアクティブ行の空欄は列単位で1件に集約・先頭セルが番地になる"""
        maps = [
            self._map("計画実績区分", "C8", "計画"),
            # 行30: 費用あり／行31: 解体機器のみ（費用空欄）／行32: 員数のみ（費用空欄）
            self._map("解体機器表_30_計画_費用", "J30", "1000000"),
            self._map("解体機器表_31_解体機器", "D31", "換気制御盤"),
            self._map("解体機器表_32_計画_員数", "G32", "3"),
        ]
        items = self._run(maps, fields=[])
        assert len(items) == 1
        item = items[0]
        assert item.field_name == "解体機器表_計画_費用"
        assert item.cell_address == "J31"
        assert "J31" in item.comment and "J32" in item.comment and "J30" not in item.comment

    def test_table_fully_filled_no_items(self):
        """アクティブ行の必須列が全て記載済みなら指摘なし（非アクティブ行は対象外）"""
        maps = [
            self._map("計画実績区分", "C8", "計画"),
            self._map("解体機器表_30_計画_費用", "J30", "1000000"),
            self._map("解体機器表_30_解体機器", "D30", "ポンプ"),
        ]
        assert self._run(maps, fields=[]) == []

    def test_table_jisseki_checks_actual_column(self):
        """実績提出なら実績_費用(N列)をチェックする"""
        maps = [
            self._map("計画実績区分", "C8", "実績"),
            self._map("解体機器表_30_解体機器", "D30", "ポンプ"),  # アクティブ行・N30空欄
        ]
        items = self._run(maps, fields=[])
        assert len(items) == 1
        assert items[0].cell_address == "N30"

    def test_no_declarations_returns_empty(self):
        """宣言が空なら何もチェックしない（opt-in）"""
        assert self._run([self._map("実施費用低減策", "G24", "")], fields=[], table={}) == []


class TestMissingEntryGolden:
    """ゴールデンExcelに対する空欄チェックの実測検証（measure-first・過検出防止）。

    完成版ゴールデンは実施理由/実施費用低減策/説明（MRC1）・年度単位総額_N年度（MRC2）が
    実際に空欄（正本データの既知の状態・2026-07-06実測）。このテストは
    「既知の空欄だけを検出し、記載済み欄には一切指摘しない」ことを固定する。
    ゴールデン側の空欄が埋められたら期待値を空集合に更新する。
    """

    _GOLDEN = "data/golden/frameB/フレームB_転記結果__ダミー完成版.xlsx"
    _KNOWN_GAPS = {
        "MRC1": {("実施理由", "G23"), ("実施費用低減策", "G24"), ("説明", "G25")},
        "MRC2": {("年度単位総額_N年度", "C30")},
    }

    def test_golden_detects_known_gaps_only(self):
        """完成版の既知の空欄のみを検出（＝それ以外への過検出ゼロ）"""
        import pathlib
        if not pathlib.Path(self._GOLDEN).exists():
            pytest.skip("ゴールデンExcelなし")
        from apps.backend.app.preliminary_review.review_logic import _generate_missing_entry_items
        from apps.backend.app.preliminary_review.criteria_loader import load_required_entries
        from apps.backend.app.preliminary_review.knowledge.result_reader import reconstruct_mappings_from_excel
        for sheet in ("MRC1", "MRC2"):
            mappings = reconstruct_mappings_from_excel(self._GOLDEN, "frameB", sheet)
            req = load_required_entries("frameB", sheet)
            items = _generate_missing_entry_items(
                mappings, "frameB", sheet,
                req["required_fields"], req["required_table_columns"],
            )
            detected = {(i.field_name, i.cell_address) for i in items}
            assert detected == self._KNOWN_GAPS[sheet], (
                f"{sheet}: 期待 {self._KNOWN_GAPS[sheet]} に対し検出 {detected}"
            )



class TestProseCitationResolution:
    """散文引用（Gemini 3.5系の形式ゆれ）の決定論解決（apply_relevance_guard）。"""

    _MAPS = [{"field_name": "対象費目1", "cell_address": "G6", "value": "施設解体一解体費"}]

    def _f3(self, doc_id, fee="解体撤去費"):
        return {"cost_category": fee, "fee_type": fee,
                "message_content": "工数・単価の積算根拠を提出されたい", "_doc_id": doc_id}

    def _item(self, evidence, src="F3"):
        from apps.backend.app.api.models import ReviewItem
        return ReviewItem(item_id="", field_name="解体機器表_30_計画_費用", cell_address="J30",
                          severity="AIからの指摘", comment="根拠不明", evidence=evidence,
                          knowledge_source=src)

    def test_prose_citation_resolved_and_normalized(self):
        """メッセージID散文引用を逆引きし、F3根拠を維持・参照番号形式へ正規化する"""
        from apps.backend.app.preliminary_review.review_logic import apply_relevance_guard
        f3_own = [self._f3("03_KT_1G_01_0003_01"), self._f3("03_KT_1G_01_0003_02")]
        it = self._item("【F3ナレッジ（自社：関東電力）｜シートKNI_1G_01｜メッセージID 03_KT_1G_01_0003_02】, "
                        "【F3ナレッジ（自社：関東電力）｜シートKNI_1G_01｜メッセージID 03_KT_1G_01_0003_01】")
        apply_relevance_guard([it], self._MAPS, [], f3_own, [])
        assert it.knowledge_source == "F3", "散文引用でもF3根拠を維持すべき"
        assert it.evidence == "[F3own#2], [F3own#1]", f"正準形式へ正規化されるべき: {it.evidence}"

    def test_prose_citation_of_irrelevant_record_still_demoted(self):
        """散文引用でも無関係レコード（費目不一致）なら従来どおり降格（難4安全性）"""
        from apps.backend.app.preliminary_review.review_logic import apply_relevance_guard
        f3_own = [self._f3("03_KT_1G_01_0009_01", fee="放射線管理費")]
        maps = [{"field_name": "対象費目1", "cell_address": "G6", "value": "処分費"}]
        it = self._item("メッセージID 03_KT_1G_01_0009_01 を参照")
        apply_relevance_guard([it], maps, [], f3_own, [])
        assert it.knowledge_source == "AI知見"

    def test_unresolvable_prose_citation_demoted(self):
        """どのレコードIDにも一致しない散文引用は従来どおり降格"""
        from apps.backend.app.preliminary_review.review_logic import apply_relevance_guard
        it = self._item("過去事例（詳細不明）による")
        apply_relevance_guard([it], self._MAPS, [], [self._f3("03_KT_1G_01_0001_01")], [])
        assert it.knowledge_source == "AI知見"

    def test_bracket_citation_path_unchanged(self):
        """従来の [F3own#N] 形式はそのまま機能（evidenceも不変）"""
        from apps.backend.app.preliminary_review.review_logic import apply_relevance_guard
        it = self._item("[F3own#1]")
        apply_relevance_guard([it], self._MAPS, [], [self._f3("03_KT_1G_01_0001_01")], [])
        assert it.knowledge_source == "F3"
        assert it.evidence == "[F3own#1]"

    def test_substring_id_collision_prefers_longer_id(self):
        """IDが部分文字列関係にある場合、長いIDを先に照合して誤解決しない"""
        from apps.backend.app.preliminary_review.review_logic import apply_relevance_guard
        f3_own = [self._f3("03_KT_0001"), self._f3("03_KT_0001_01")]
        it = self._item("メッセージID 03_KT_0001_01 参照")
        apply_relevance_guard([it], self._MAPS, [], f3_own, [])
        assert it.evidence == "[F3own#2]", f"長いID(#2)に解決されるべき: {it.evidence}"

    def test_mixed_prose_and_bracket_normalized(self):
        """散文とbracketが混在するevidenceも出現順で正準形式に正規化"""
        from apps.backend.app.preliminary_review.review_logic import apply_relevance_guard
        f3_own = [self._f3("03_KT_1G_01_0003_01"), self._f3("03_KT_1G_01_0003_02")]
        it = self._item("【F3ナレッジ｜メッセージID 03_KT_1G_01_0003_02】 および [F3own#1] を参照")
        apply_relevance_guard([it], self._MAPS, [], f3_own, [])
        assert it.knowledge_source == "F3"
        assert it.evidence == "[F3own#2], [F3own#1]", f"混在引用も正規化: {it.evidence}"

    def test_no_submission_fee_normalizes_but_never_demotes(self):
        """費目が取れないシート（MRC2等）でも正規化は行い、降格はしない（従来挙動維持）"""
        from apps.backend.app.preliminary_review.review_logic import apply_relevance_guard
        maps_no_fee = [{"field_name": "費用内訳_人件費_数量", "cell_address": "E39", "value": "1"}]
        f3_own = [self._f3("03_KT_1G_01_0004_01", fee="全く別の費目")]
        it = self._item("【F3ナレッジ｜メッセージID 03_KT_1G_01_0004_01】")
        apply_relevance_guard([it], maps_no_fee, [], f3_own, [])
        assert it.knowledge_source == "F3", "費目不明時は降格しない（従来どおりガードオフ）"
        assert it.evidence == "[F3own#1]", f"正規化はされる: {it.evidence}"


class TestMergeRuleAndGeminiItems:
    """ルール指摘×Gemini指摘のマージ（F2/F3根拠つきはルールより優先・1セル1指摘）。"""

    def _item(self, cell, src, comment="c"):
        from apps.backend.app.api.models import ReviewItem
        return ReviewItem(item_id="", field_name="f", cell_address=cell,
                          severity="要確認", comment=comment, evidence="", knowledge_source=src)

    def test_grounded_gemini_wins_over_rule_on_same_cell(self):
        """同一セルではF3根拠つきGemini指摘がルール指摘（必須欄空欄等）に勝つ"""
        from apps.backend.app.preliminary_review.review_logic import merge_rule_and_gemini_items
        rule   = [self._item("G24", "AI知見", "記載必須欄が空欄")]
        gemini = [self._item("G24", "F3", "過去事例に基づく指摘")]
        out = merge_rule_and_gemini_items(rule, gemini)
        assert len(out) == 1 and out[0].knowledge_source == "F3"

    def test_ungrounded_gemini_loses_to_rule_on_same_cell(self):
        """根拠なし（AI知見）のGemini指摘は従来どおりルール指摘を優先"""
        from apps.backend.app.preliminary_review.review_logic import merge_rule_and_gemini_items
        rule   = [self._item("G24", "AI知見", "記載必須欄が空欄")]
        gemini = [self._item("G24", "AI知見", "LLMの重複指摘")]
        out = merge_rule_and_gemini_items(rule, gemini)
        assert len(out) == 1 and out[0].comment == "記載必須欄が空欄"

    def test_disjoint_cells_all_kept(self):
        """セルが重ならなければ両方残る（ルール→Geminiの順）"""
        from apps.backend.app.preliminary_review.review_logic import merge_rule_and_gemini_items
        rule   = [self._item("G25", "AI知見")]
        gemini = [self._item("J30", "F3"), self._item("G23", "AI知見")]
        out = merge_rule_and_gemini_items(rule, gemini)
        assert [i.cell_address for i in out] == ["G25", "J30", "G23"]
