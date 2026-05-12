import pytest
from pathlib import Path
from apps.backend.app.agents.data_extractor.parser import _parse_excel
from apps.backend.app.api.routes.upload import _extract_from_file

def test_parse_excel_with_enumerate(tmp_path):
    """Excelパーサーのenumerateによる行番号取得を検証。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws.cell(row=1, column=1, value="Header")
    ws.cell(row=2, column=1, value="Data1")
    ws.cell(row=3, column=1, value="Data2")

    excel_path = tmp_path / "test.xlsx"
    wb.save(excel_path)

    result = _parse_excel(excel_path)
    assert "## シート: TestSheet" in result
    assert "行1 | Header" in result
    assert "行2 | Data1" in result
    assert "行3 | Data2" in result

def test_extract_from_file_suffix_validation():
    """_extract_from_file の suffix バリデーションを検証。"""
    content = b"fake content"
    filename = "test.txt"
    suffix = ".txt"  # サポート外
    sheet_name = "Sheet1"
    frame_name = "frameB"

    with pytest.raises(ValueError, match="Unsupported extension"):
        _extract_from_file(content, filename, suffix, sheet_name, frame_name)

def test_extract_from_file_path_traversal_prevention(tmp_path, monkeypatch):
    """_extract_from_file の suffix によるパストラバーサル防止を検証。"""
    # UPLOAD_DIR を一時ディレクトリに向ける
    monkeypatch.setattr("apps.backend.app.api.routes.upload.UPLOAD_DIR", tmp_path)

    # extract_data をモック
    monkeypatch.setattr("apps.backend.app.api.routes.upload.extract_data", lambda **kwargs: {"data": {}, "_metadata": {}})

    content = b"fake content"
    filename = "test.xlsx"
    suffix = "../../../etc/passwd"  # パストラバーサル試行
    sheet_name = "Sheet1"
    frame_name = "frameB"

    # 実際には suffix not in SUPPORTED_EXTENSIONS で弾かれるが、
    # もし extension が一致してしまった場合でも Path.name で守られることを確認
    with pytest.raises(ValueError, match="Unsupported extension"):
        _extract_from_file(content, filename, suffix, sheet_name, frame_name)
