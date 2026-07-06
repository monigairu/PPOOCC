"""
template.py の表示用レイアウト生成・簡易式評価のテスト

実行方法:
    PYTHONPATH=. uv run pytest apps/backend/tests/test_template_layout.py -v
"""
import pytest
from openpyxl import Workbook

from apps.backend.app.api.routes.template import _read_excel_layout


@pytest.fixture
def workbook_path(tmp_path):
    """対応式・未対応式・異常式を1シートに詰めたExcelを作る。

    Returns:
        式評価の代表ケースを含む一時Excelファイルのパス。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "MRC1"
    ws["A1"] = 100
    ws["A2"] = 200
    ws["A3"] = "=SUM(A1:A2)"                     # 範囲SUM
    ws["A4"] = "=A3*2"                            # 式セル参照＋乗算
    ws["A5"] = '=IF(A1=100,"百","他")'            # IF等価条件
    ws["A6"] = '="値:"&A1'                        # 文字列連結
    ws["B1"] = "=B2"                              # 循環参照
    ws["B2"] = "=B1"
    ws["B3"] = "=VLOOKUP(A1,A1:A2,1)"             # 未対応関数
    ws["B4"] = "=Sheet9!A1"                       # 存在しないシート参照
    ws["B5"] = "=SUM(Sheet9!A1:A2)"               # SUM内の存在しないシート範囲
    ws["B6"] = "=MRC2!A1"                         # 存在するクロスシート参照
    ws2 = wb.create_sheet("MRC2")
    ws2["A1"] = 999
    path = tmp_path / "layout_test.xlsx"
    wb.save(path)
    return str(path)


def _cell_map(layout):
    return {c["address"]: c for c in layout["cells"]}


def test_supported_formulas_evaluated(workbook_path):
    """SUM・参照演算・IF・連結・クロスシートが表示値へ評価される"""
    cells = _cell_map(_read_excel_layout(workbook_path, "MRC1"))
    assert cells["A3"]["value"] == "300"
    assert cells["A4"]["value"] == "600"
    assert cells["A5"]["value"] == "百"
    assert cells["A6"]["value"] == "値:100"
    assert cells["B6"]["value"] == "999"


def test_formula_string_preserved(workbook_path):
    """評価可否によらず formula フィールドに式文字列が残る"""
    cells = _cell_map(_read_excel_layout(workbook_path, "MRC1"))
    assert cells["A3"]["formula"] == "=SUM(A1:A2)"
    assert cells["B3"]["formula"] == "=VLOOKUP(A1,A1:A2,1)"
    assert cells["A1"]["formula"] is None  # 通常値セルは None


def test_missing_sheet_reference_does_not_crash(workbook_path):
    """存在しないシート参照の式は None フォールバックし、API全体を落とさない（KeyError→500の回帰防止）"""
    cells = _cell_map(_read_excel_layout(workbook_path, "MRC1"))
    assert cells["B4"]["value"] is None
    assert cells["B5"]["value"] is None


def test_unsupported_and_circular_fallback_to_none(workbook_path):
    """未対応関数・循環参照は例外にせず表示値 None"""
    cells = _cell_map(_read_excel_layout(workbook_path, "MRC1"))
    assert cells["B3"]["value"] is None
    assert cells["B1"]["value"] is None
    assert cells["B2"]["value"] is None
