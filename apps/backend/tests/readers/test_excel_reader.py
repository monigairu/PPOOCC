import pytest
from pathlib import Path
from openpyxl import Workbook


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """テスト用のシンプルな Excel ファイルを生成する"""
    wb = Workbook()
    ws = wb.active
    ws.title = "物量リスト"
    ws.append(["機器ID", "機器名", "口径(A)", "重量(t)"])
    ws.append(["P-001", "配管A", 50, 1.5])
    ws.append(["P-002", "配管B", 100, 2.8])

    ws2 = wb.create_sheet("参考")
    ws2.append(["備考"])
    ws2.append(["サンプルデータ"])

    path = tmp_path / "物量データ_sample.xlsx"
    wb.save(path)
    return path


def test_read_excel_returns_source_document(sample_xlsx):
    from apps.backend.app.readers.excel_reader import read_excel
    doc = read_excel(str(sample_xlsx))

    assert doc.source_type == "excel"
    assert doc.source_file == str(sample_xlsx)


def test_read_excel_document_kind_inferred(sample_xlsx):
    from apps.backend.app.readers.excel_reader import read_excel
    doc = read_excel(str(sample_xlsx))
    # ファイル名に "物量" を含むので "物量データ" になる
    assert doc.document_kind == "物量データ"


def test_read_excel_text_content_contains_sheet_name(sample_xlsx):
    from apps.backend.app.readers.excel_reader import read_excel
    doc = read_excel(str(sample_xlsx))
    assert "物量リスト" in doc.text_content


def test_read_excel_text_content_contains_cell_values(sample_xlsx):
    from apps.backend.app.readers.excel_reader import read_excel
    doc = read_excel(str(sample_xlsx))
    assert "配管A" in doc.text_content
    assert "P-001" in doc.text_content


def test_read_excel_metadata_contains_sheets(sample_xlsx):
    from apps.backend.app.readers.excel_reader import read_excel
    doc = read_excel(str(sample_xlsx))
    assert "物量リスト" in doc.metadata["sheets"]
    assert "参考" in doc.metadata["sheets"]


def test_read_excel_all_sheets_dumped(sample_xlsx):
    from apps.backend.app.readers.excel_reader import read_excel
    doc = read_excel(str(sample_xlsx))
    # 2枚目シートの内容も含まれているか
    assert "サンプルデータ" in doc.text_content


def test_read_excel_unknown_kind_for_unnamed_file(tmp_path):
    from apps.backend.app.readers.excel_reader import read_excel
    wb = Workbook()
    wb.active.append(["test"])
    path = tmp_path / "AA_20250401.xlsx"
    wb.save(path)

    doc = read_excel(str(path))
    assert doc.document_kind == "不明"
