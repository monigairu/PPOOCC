"""
Excel リーダー

既存の parser.py:_parse_excel を再利用して SourceDocument を構築する。
Excel 読み込みロジックを二重実装しない。
"""
from openpyxl import load_workbook

from apps.backend.app.agents.data_extractor.parser import parse_file
from .source_document import SourceDocument, infer_document_kind


def read_excel(file_path: str) -> SourceDocument:
    """
    Excel ファイルを読み込んで SourceDocument を返す。

    テキスト変換は parser.py に委譲する（全シート・行番号プレフィックス付きダンプ）。
    """
    text_content = parse_file(file_path)

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    wb.close()

    return SourceDocument(
        source_file=file_path,
        source_type="excel",
        document_kind=infer_document_kind(file_path),
        text_content=text_content,
        metadata={"sheets": sheet_names},
    )
