"""
F2/F3ナレッジ Excel 読み込み専用モジュール

knowledge_loader.py（検索バックエンド）とは独立して使用する。
ingest_knowledge.py などのデータ投入スクリプトから直接呼び出す。

Phase が変わっても Excel をソースとして読み込む処理はここで管理する。
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_KNOWLEDGE_DIR = Path("data/knowledge")
_SCHEMA_DIR = _KNOWLEDGE_DIR / "schema"

# ver5.3 平坦形式（1メッセージ=1行）の正準列（REQUIREMENTS.md §0-7 R1）。
# BigQuery 平坦テーブル・Agent Search 索引のスキーマ契約として使う。
# 各キーは schema YAML の fixed_columns キーおよび flatten_qa の生成キーと一致させる。
#
# F3（電力別ナレッジ・新様式 ver5.3）
F3_VER53_COLUMNS = (
    "id",                 # ID
    "message_id",         # メッセージID（flatten_qa が生成: {id}_{seq:02d}）
    "start_date",         # 起票日
    "dept_group",         # 起票者所属G
    "author",             # 起票者
    "ref_knowledge_id",   # 参照先ナレッジID
    "submission_timing",  # 提出タイミング
    "confirm_year",       # 確認年度
    "plant_site",         # 該当発電所
    "plant_unit",         # 該当プラント
    "cost_category",      # 該当費目
    "construction_name",  # 該当工事
    "reference_url",      # 該当資料
    "message_content",    # メッセージ内容（検索対象テキスト・flatten_qa が生成）
)

# F3 本体列に加えて検索フィルタ・権限制御に使う付帯列（RAG_VERIFICATION.md §3-3）
F3_VER53_AUX_COLUMNS = (
    "utility_name",       # 電力会社（正規化前。正規化は ingest 側で適用）
    "reactor_type",       # 炉型（BWR/PWR）。様式に列は無く該当発電所から導出（plant_reactor_map.yaml）
    "sheet_name",         # 由来シート（KNI_1G_01 等＝提出タイミング分岐の後ろ盾）
    "message_direction",  # nuro / denryoku（flatten_qa が生成）
    "round",              # やりとり回数（flatten_qa が生成）
)

# F2（NuRO内共有の問合せ・知見ナレッジ・新様式 ver5.3）。画像の出力用シート列順に準拠。
F2_VER53_COLUMNS = (
    "id",                  # ID
    "message_id",          # メッセージID（flatten_qa が生成）
    "start_date",          # 起票日
    "category",            # 区分
    "dept_group",          # 起票者所属G
    "author",              # 起票者
    "business_category",   # 業務カテゴリ
    "related_group",       # 関連グループ名
    "ref_knowledge_id",    # 参照先ナレッジID
    "from_party",          # From
    "to_party",            # To
    "reference_url",       # 該当資料
    "priority",            # 優先度
    "status",              # ステータス
    "title",               # タイトル
    "contact_medium",      # 連絡媒体
    "related_material_1",  # 関連資料①
    "related_material_2",  # 関連資料②
    "related_material_3",  # 関連資料③
    "phenomenon_summary",  # 事象概要
    "judgment_basis",      # 判断基準
    "response_result",     # 対応結果
    "message_content",     # メッセージ内容（検索対象テキスト・flatten_qa が生成）
)

# F2 付帯列。F2 は NuRO 内共有のため電力会社・炉型は持たない。
# caller_role_required は権限フィルタ（load_f2 が caller_role_required=NuRO で絞る）用の定数列。
F2_VER53_AUX_COLUMNS = (
    "sheet_name",           # 由来シート（KNS_1G 等）
    "message_direction",    # question / answer（flatten_qa が生成）
    "round",                # やりとり回数（flatten_qa が生成）
    "caller_role_required", # 権限フィルタ用（F2 は常に "NuRO"・ingest で付与）
)

# knowledge_type → (本体列, 付帯列) のレジストリ。to_ver53_rows / ingest / 索引が参照する単一の契約。
VER53_SCHEMA: dict[str, tuple[tuple[str, ...], tuple[str, ...]]] = {
    "f3": (F3_VER53_COLUMNS, F3_VER53_AUX_COLUMNS),
    "f2": (F2_VER53_COLUMNS, F2_VER53_AUX_COLUMNS),
}

# 発電所→炉型の導出マップ（ドメイン知識・config）。様式のZ列は廃止（2026-07-02）
_PLANT_REACTOR_MAP_PATH = _SCHEMA_DIR / "plant_reactor_map.yaml"
_plant_reactor_map: dict[str, str] | None = None


def _load_plant_reactor_map() -> dict[str, str]:
    global _plant_reactor_map
    if _plant_reactor_map is None:
        if _PLANT_REACTOR_MAP_PATH.exists():
            with open(_PLANT_REACTOR_MAP_PATH, encoding="utf-8") as f:
                _plant_reactor_map = yaml.safe_load(f) or {}
        else:
            _plant_reactor_map = {}
    return _plant_reactor_map


def derive_reactor_type(plant_site: str, plant_unit: str = "") -> str:
    """該当発電所（＋号機）から炉型を導出する。

    正本Excel（ver5.3様式）に炉型の列は存在しないため、発電所→炉型の
    ドメイン知識（plant_reactor_map.yaml）から導出する。
    号機で炉型が異なる発電所は「発電所名/号機」キーが発電所名キーより優先。
    マップに無い発電所は ""（炉型不明＝フィルタ対象外）。
    """
    if not plant_site:
        return ""
    reactor_map = _load_plant_reactor_map()
    if plant_unit and f"{plant_site}/{plant_unit}" in reactor_map:
        return str(reactor_map[f"{plant_site}/{plant_unit}"])
    return str(reactor_map.get(plant_site, ""))


def to_ver53_rows(
    records: list[dict[str, Any]],
    knowledge_type: str = "f3",
) -> list[dict[str, Any]]:
    """
    read_all_f2()/read_all_f3() の平坦レコードを ver5.3 の正準列＋付帯列に射影する。

    - knowledge_type（"f2"/"f3"）に対応する列だけを残す（余分なキーは落とす）
    - 欠けているキーは "" で埋める（round のみ 0）
    → BigQuery ロード・Agent Search 索引が常に同一スキーマの行を受け取れる。
    """
    columns, aux_columns = VER53_SCHEMA[knowledge_type]
    rows: list[dict[str, Any]] = []
    for record in records:
        row: dict[str, Any] = {}
        for key in columns + aux_columns:
            default: Any = 0 if key == "round" else ""
            row[key] = record.get(key, default)
        rows.append(row)
    return rows


def _apply_bq_field_defaults(knowledge_type: str, rows: list[dict[str, Any]]) -> None:
    """knowledge_type ごとの ingest 時フィールド補正（検索側と一貫させる）。"""
    # 循環import回避のため関数内import（knowledge_loader は検索バックエンドに依存する）
    from apps.backend.app.agents.reviewer.knowledge_loader import normalize_utility

    if knowledge_type == "f3":
        # 会社名は検索側（load_f3）・直接投入（_build_document）と同じ正規化で表記ゆれを吸収
        for row in rows:
            row["utility_name"] = normalize_utility(row["utility_name"])
    elif knowledge_type == "f2":
        # F2 は NuRO のみ参照可。load_f2 が caller_role_required=NuRO で絞るため定数を付与
        for row in rows:
            row["caller_role_required"] = "NuRO"


def excel_to_bq_input(
    records: list[dict[str, Any]],
    knowledge_type: str,
) -> list[dict[str, Any]]:
    """Excel読み取り結果を受け取って BigQuery のインプットに加工する親玉関数。

    read_all_f2()/read_all_f3() が返した平坦レコード（1メッセージ=1行）を、
    BigQuery ロード（→ Agent Search 索引）へそのまま渡せる行リストに変換する。

    中身は次の3段（各処理の詳細はそれぞれの関数を参照）:
      1. to_ver53_rows()            正準列契約 VER53_SCHEMA への射影
      2. _apply_bq_field_defaults() knowledge_type ごとの検索用フィールド補正
      3. message_id 検証            空の行があれば ValueError（索引の id_field に使えない）

    Args:
        records: read_all_f2()/read_all_f3() の戻り値（平坦レコードのリスト）
        knowledge_type: "f2" または "f3"

    Returns:
        BigQuery へロード可能な ver5.3 行（dict）のリスト。
        例: [{"id": "F3-001", "message_id": "F3-001_01", "message_content": "...", ...}, ...]

    Raises:
        ValueError: message_id が空の行がある場合（件数をメッセージに含む）

    使用例:
        records = read_all_f3()
        rows = excel_to_bq_input(records, "f3")   # → そのまま BigQuery へロードできる
    """
    assert isinstance(records, list), "records は read_all_f2()/read_all_f3() の戻り値（list）を渡す"
    assert knowledge_type in ("f2", "f3"), f"knowledge_type は 'f2' か 'f3'（実際: {knowledge_type!r}）"

    rows = to_ver53_rows(records, knowledge_type)
    _apply_bq_field_defaults(knowledge_type, rows)

    no_id = [r for r in rows if not r["message_id"]]
    if no_id:
        raise ValueError(
            f"message_id が空の行が {len(no_id)} 件あります（flatten_qa=False のスキーマ？）。"
            "Agent Search 索引の id_field に使えないため中止します"
        )
    return rows


def _col_letter_to_idx(col: str) -> int:
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def _infer_direction(key: str) -> str:
    k = key.lower()
    if "nuro" in k or "question" in k:
        return "nuro"
    if "denryoku" in k or "reply" in k or "answer" in k:
        return "denryoku"
    return "unknown"


def _discover_schemas(frame_prefix: str) -> list[dict]:
    if not _SCHEMA_DIR.exists():
        return []
    schemas = []
    for path in sorted(_SCHEMA_DIR.glob(f"{frame_prefix}_*_schema.yaml")):
        try:
            with open(path, encoding="utf-8") as f:
                schemas.append(yaml.safe_load(f))
        except Exception as e:
            logger.warning("スキーマ読み込みエラー: %s (%s)", path, e)
    return schemas


def _expand_merged_cells(df_raw: pd.DataFrame, ws) -> None:
    """結合セルの値を結合範囲全体に展開する（df_raw を直接書き換える）。

    Excelの結合セルは左上（アンカー）セルにしか値を持たず、グリッド化すると
    残りのセルが空になる。schema の `merge_cell_handling: fill_down` の意図
    （結合された行にも同じ値を行き渡らせる）を、openpyxl の結合範囲情報を
    使って**結合セルに限定して**実現する。

    ※ 旧実装の全列 ffill（無条件の下方向前埋め）は、結合と無関係な
    「意図的に空のセル」に上の行の値を染み出させ、存在しないメッセージを
    捏造していたため廃止（2026-07-04・RAG_VERIFICATION §1-16）。

    Args:
        df_raw: ワークシートを文字列化したグリッド（行0=Excel1行目・空セルは ""）。
            この DataFrame を in-place で書き換える。
        ws: 同じシートの openpyxl ワークシート（結合範囲 merged_cells の取得元）。

    Returns:
        None（df_raw を直接更新する）。
    """
    n_rows, n_cols = df_raw.shape
    for rng in ws.merged_cells.ranges:
        r0, c0 = rng.min_row - 1, rng.min_col - 1  # 0始まりに変換
        if r0 >= n_rows or c0 >= n_cols:
            continue
        anchor = df_raw.iat[r0, c0]
        for r in range(r0, min(rng.max_row, n_rows)):
            for c in range(c0, min(rng.max_col, n_cols)):
                df_raw.iat[r, c] = anchor


def _read_excel_by_schema(
    schema: dict,
    file_path: Path,
) -> tuple[list[dict[str, Any]], str]:
    layout = schema.get("layout", {})
    data_start_row: int = layout.get("data_start_row", 7)
    loader_cfg: dict = schema.get("loader_config", {})
    id_col: str = loader_cfg.get("id_column", "A")
    id_col_idx: int = _col_letter_to_idx(id_col)

    # openpyxl 1回読みでグリッド化（旧: pd.read_excel＋結合範囲取得のための再読込＝二重ロード）。
    # 値は文字列化して空セルは ""（旧 dtype=str・fillna("") と同等の契約を維持）
    excel_sheet = schema.get("excel_sheet")
    wb = load_workbook(file_path, data_only=True)
    ws = wb[excel_sheet] if excel_sheet and excel_sheet in wb.sheetnames else wb.worksheets[0]
    df_raw = pd.DataFrame([
        ["" if v is None else str(v) for v in row]
        for row in ws.iter_rows(values_only=True)
    ])
    # 結合セルのみアンカー値を展開（merge_cell_handling: fill_down の正しい実装）
    _expand_merged_cells(df_raw, ws)

    utility_name = ""
    for _meta_key, meta_val in schema.get("meta_cells", {}).items():
        cell_addr: str = meta_val.get("cell", "")
        if not cell_addr:
            continue
        col_letter = "".join(c for c in cell_addr if c.isalpha())
        row_num = int("".join(c for c in cell_addr if c.isdigit()))
        c_idx = _col_letter_to_idx(col_letter)
        r_idx = row_num - 1
        if r_idx < len(df_raw) and c_idx < df_raw.shape[1]:
            utility_name = str(df_raw.iat[r_idx, c_idx]).strip()
        break

    if data_start_row - 1 >= len(df_raw):
        return [], utility_name

    # 旧実装はここで全列を無条件 ffill していたが、結合と無関係な空セルに
    # 上の行の値が染み出し、メッセージの捏造・固定列の混入を起こすため廃止。
    # 結合セルの展開は _expand_merged_cells() で対応済み。空セルは空のまま扱う。
    data_df = df_raw.iloc[data_start_row - 1:].copy().reset_index(drop=True)

    fixed_columns: list[dict] = schema.get("fixed_columns", [])
    qa_config: dict | None = schema.get("repeating_qa_columns")
    flatten_qa: bool = schema.get("output_model", {}).get("flatten_qa", True)
    records: list[dict[str, Any]] = []

    for _, row in data_df.iterrows():
        id_val = row.iloc[id_col_idx].strip() if id_col_idx < len(row) else ""
        if not id_val or id_val in ("nan", "None"):
            continue

        base: dict[str, Any] = {}
        for col_def in fixed_columns:
            c_idx = _col_letter_to_idx(col_def["col"])
            val = row.iloc[c_idx].strip() if c_idx < len(row) else ""
            base[col_def["key"]] = "" if val in ("nan", "None") else val

        if utility_name:
            base["utility_name"] = utility_name

        # 由来シート（KNI_1G_01 等）。提出タイミング分岐・BigQuery平坦テーブルで使う
        if schema.get("sheet_name"):
            base["sheet_name"] = schema["sheet_name"]

        # 炉型は様式の列ではなく該当発電所から導出（plant_reactor_map.yaml・ドメイン知識）
        if base.get("plant_site") and not base.get("reactor_type"):
            derived_rt = derive_reactor_type(base["plant_site"], base.get("plant_unit", ""))
            if derived_rt:
                base["reactor_type"] = derived_rt

        if qa_config and flatten_qa:
            start_col_idx = _col_letter_to_idx(qa_config["start_col"])
            col_per_round: int = qa_config["col_per_round"]
            max_rounds: int = qa_config["max_rounds"]
            qa_fields: list[dict] = qa_config["fields"]

            # message_id は ID ごとに横に連なるメッセージの通し連番（公式 ver5.3 出力用シートに準拠）。
            # 読み順（round 昇順 → field 順）で非空メッセージを 01, 02, … と採番する。
            # round と message_direction は別カラムとして保持するため情報は失われない
            # （旧独自形式 {id}_{round}_{direction} を公式の通し連番に統一）。
            seq = 0
            for round_num in range(1, max_rounds + 1):
                for field_def in qa_fields:
                    actual_idx = (
                        start_col_idx
                        + (round_num - 1) * col_per_round
                        + field_def["col_offset"]
                    )
                    if actual_idx >= len(row):
                        continue
                    content = row.iloc[actual_idx].strip()
                    if not content or content in ("nan", "None"):
                        continue
                    seq += 1
                    msg_record = {**base}
                    msg_record["message_id"] = f"{id_val}_{seq:02d}"
                    msg_record["round"] = round_num
                    msg_record["message_direction"] = _infer_direction(field_def["key"])
                    msg_record["message_content"] = content
                    records.append(msg_record)
        else:
            records.append(base)

    return records, utility_name


def read_all_f2() -> list[dict[str, Any]]:
    """F2ナレッジをExcelから全件読み込む。"""
    schemas = _discover_schemas("f2")
    all_records: list[dict] = []
    for schema in schemas:
        frame = schema.get("frame", "F2").upper()
        sname = schema.get("sheet_name", "")
        excel_file = schema.get("excel_file") or f"{frame}_{sname}.xlsx"
        file_path = _KNOWLEDGE_DIR / excel_file
        if not file_path.exists():
            continue
        try:
            records, _ = _read_excel_by_schema(schema, file_path)
            all_records.extend(records)
        except Exception as e:
            logger.warning("F2読み込みエラー: %s (%s)", file_path, e)
    return all_records


def read_all_f3() -> list[dict[str, Any]]:
    """F3ナレッジをExcelから全件読み込む。"""
    schemas = _discover_schemas("f3")
    all_records: list[dict] = []
    for schema in schemas:
        frame = schema.get("frame", "F3").upper()
        sname = schema.get("sheet_name", "")
        excel_file = schema.get("excel_file") or f"{frame}_{sname}.xlsx"
        file_path = _KNOWLEDGE_DIR / excel_file
        if not file_path.exists():
            continue
        try:
            records, _ = _read_excel_by_schema(schema, file_path)
            all_records.extend(records)
        except Exception as e:
            logger.warning("F3読み込みエラー: %s (%s)", file_path, e)
    return all_records
