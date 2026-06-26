"""
転記結果Excel → mappings 復元モジュール

様式定義（config の セル↔フィールド 定義）を使って、転記済みの「結果Excel」から
レビュー入力となる mappings（{field_name, cell_address, value, reasoning}）を復元する。
cell_writer の逆操作にあたる。

利用元:
  - scripts/verify_rag.py（RAG検証ハーネス）
  - api/routes/review.py（様式Excelをアップロードしてレビューする入口）

knowledge_loader（検索バックエンド）・_excel_reader（F2/F3ナレッジ読み込み）とは独立。
I/F は不変を保つ（様式定義に依存し、特定費目・特定ファイルをハードコードしない）。
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from apps.backend.app.core.frame_config_loader import (
    load_frame_config,
    extract_cell_definitions,
)


def reconstruct_mappings_from_excel(
    excel_path: Path | str, frame: str, sheet: str
) -> list[dict]:
    """frame config の セル↔フィールド 定義を使って結果Excelから mappings を復元する。

    {field_name, cell_address, value, reasoning} のリストを返す。
    同一フィールドが複数セル（計画/実績）を持つ場合は値のあるセルのみ採用する。
    """
    config = load_frame_config(frame, sheet)
    cell_defs = extract_cell_definitions(config)  # label_value / plan_actual のみ

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active

    mappings: list[dict] = []
    seen: set[tuple[str, str]] = set()

    def _add(field_name: str, cell: str) -> None:
        raw = ws[cell].value
        if raw is None or str(raw).strip() == "":
            return
        key = (field_name, cell)
        if key in seen:
            return
        seen.add(key)
        mappings.append({
            "field_name":   field_name,
            "cell_address": cell,
            "value":        str(raw).strip(),
            "reasoning":    "",
        })

    # label_value / plan_actual（既存ユーティリティが拾う分）
    for field_name, cells in cell_defs.items():
        for cell in cells:
            _add(field_name, cell)

    # tabular（解体機器表・費用内訳など）: extract_cell_definitions が拾わないため個別処理
    for section in config.get("sections", []):
        if section.get("type") != "tabular":
            continue
        _add_tabular(section, ws, _add)

    return mappings


def _add_tabular(section: dict, ws, add) -> None:
    """tabular セクションのセルを mappings に追加する。

    2形式に対応:
      ① row_match.rows で行を明示（例: MRC2 費用内訳の 人件費/材料費…）
      ② data_start_row から空行まで動的スキャン（例: MRC1 解体機器表）
    field_name は「{セクション名}_{行ID}_{列名}」。
    """
    sec_name = section.get("name", "表")
    cols = [c for c in section.get("columns", []) if c.get("column")]
    if not cols:
        return

    explicit_rows = section.get("row_match", {}).get("rows")
    if explicit_rows:
        rows = [(rm["row"], str(rm.get("row_id", rm["row"]))) for rm in explicit_rows if rm.get("row")]
    else:
        start = section.get("data_start_row")
        if not start:
            return
        end = section.get("data_end_row") or (start + 200)  # 上限ガード
        rows = []
        for r in range(int(start), int(end) + 1):
            # 全列が空になった行で打ち切り（動的表の終端検出）
            if all(ws[f"{c['column']}{r}"].value in (None, "") for c in cols):
                break
            rows.append((r, str(r)))

    for row, row_id in rows:
        for c in cols:
            add(f"{sec_name}_{row_id}_{c.get('name', c['column'])}", f"{c['column']}{row}")


def _derive_queries(mappings: list[dict]) -> dict[str, str | None]:
    """run_review と同じロジックで RAG クエリ用フィールドを取り出す。"""
    # 循環インポート回避のため遅延 import
    from apps.backend.app.agents.reviewer import reviewer_agent
    return {
        "fee_type":     reviewer_agent._extract_field(mappings, "対象費目1"),
        "reactor_type": reviewer_agent._extract_field(mappings, "炉型"),
        "utility_name": reviewer_agent._extract_field(mappings, "電力会社"),
    }


def derive_query_context(
    excel_path: Path | str, frame: str, target_sheet: str, context_sheet: str = "MRC1"
) -> dict[str, str | None]:
    """RAG クエリ文脈（費目・炉型・会社）を申請の基本情報シートから取得する。

    MRC2 など費目・炉型を持たないシートをレビューする場合でも、同一申請の
    基本情報シート（既定 MRC1）から費目・炉型・会社を引いてクエリに使う。
    対象シート自身に基本情報があればそちらを優先する。
    """
    q = _derive_queries(reconstruct_mappings_from_excel(excel_path, frame, target_sheet))
    if q.get("fee_type") and q.get("utility_name"):
        return q
    # 不足分を context_sheet（基本情報シート）から補完
    try:
        ctx = _derive_queries(reconstruct_mappings_from_excel(excel_path, frame, context_sheet))
    except FileNotFoundError:
        return q
    for k in ("fee_type", "reactor_type", "utility_name"):
        if not q.get(k):
            q[k] = ctx.get(k)
    return q
