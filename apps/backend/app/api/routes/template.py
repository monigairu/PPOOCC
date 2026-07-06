"""Excelテンプレートや転記済み成果物を、フロントエンド表示用レイアウトに変換する API route。

このモジュールは、openpyxl で Excel ファイルを読み、グリッド表示に必要なセル値、式、結合セル、
列幅、行高さを JSON に変換する。起動直後は空テンプレートを返し、転記後は session ごとの
出力 Excel を返す。

Excel の式は、openpyxl だけでは再計算されない。そのため、画面で必要な範囲に限り、`SUM`、
`ROUND`、`IF`、`IFERROR`、四則演算、文字列連結、セル参照、範囲参照を簡易評価する。
VLOOKUP など未対応の式は例外で落とさず、表示値を `None` のまま返す。

この route は表示用 JSON の生成だけを担当する。転記処理、Excel への書き込み、GCS アップロード、
式の完全互換評価は行わない。
"""

from pathlib import Path
import re
from fastapi import APIRouter, HTTPException, Path as FastAPIPath, Query
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils import column_index_from_string

from apps.backend.app.core.settings import TEMPLATE_PATH, OUTPUT_DIR

router = APIRouter()


_FUNCTION_RE = re.compile(r"^(?P<name>[A-Z][A-Z0-9]*)\((?P<args>.*)\)$", re.IGNORECASE)
_NUMBER_RE = re.compile(r"^[+-]?\d+(?:\.\d+)?$")
_CELL_REF_RE = re.compile(r"^(?:(?P<sheet>'[^']+'|[A-Za-z0-9_]+)!)?(?P<cell>\$?[A-Z]+\$?\d+)$")
_RANGE_REF_RE = re.compile(
    r"^(?:(?P<sheet>'[^']+'|[A-Za-z0-9_]+)!)?(?P<start>\$?[A-Z]+\$?\d+):(?P<end>\$?[A-Z]+\$?\d+)$"
)


class FormulaEvaluationError(ValueError):
    """表示用の簡易Excel式評価で扱えない式を表す例外。

    `_evaluate_expression()` が未対応関数、未対応演算、循環参照、数値変換不能などを検知したときに使う。
    上位の `_evaluate_supported_formula()` はこの例外を捕捉し、API表示では `None` にフォールバックする。

    Examples:
        >>> isinstance(FormulaEvaluationError("unsupported"), ValueError)
        True

    Note:
        この例外はユーザーへ直接返すためのものではない。未対応式を安全に表示スキップするための内部信号。
    """


@router.get("/template")
async def get_template_structure(
    sheet_name: str = Query("MRC1", pattern=r"^[a-zA-Z0-9_\-]+$")
):
    """空テンプレート Excel の表示レイアウトを返す。

    フロントエンド起動時、まだ転記済み session が無い状態でグリッドの初期表示を作るための endpoint。
    `TEMPLATE_PATH` の Excel を読み、指定 sheet のセル値・式・結合セル・列幅・行高さを
    `_read_excel_layout()` で JSON 化して返す。

    Args:
        sheet_name (str): 読み取る sheet 名。既定は `MRC1`。FastAPI Query の pattern で
            英数字・アンダースコア・ハイフンに制限する。

    Returns:
        dict: グリッド表示用 layout。主な key は `sheet_name`、`max_row`、`max_col`、
        `cells`、`merged_cells`、`col_widths`、`row_heights`。

    Raises:
        HTTPException: `TEMPLATE_PATH` が存在しない場合は 404。
        HTTPException: 指定 sheet が Excel 内に無い場合は `_read_excel_layout()` から 404 が伝播する。

    Examples:
        FastAPI endpoint と実Excelに依存するため doctest では実行しない。

        >>> response = client.get("/api/template?sheet_name=MRC1")  # doctest: +SKIP
        >>> response.json()["sheet_name"]  # doctest: +SKIP
        'MRC1'

    Note:
        この endpoint は読み取り専用。テンプレート Excel の内容は変更しない。
    """
    if not TEMPLATE_PATH.exists():
        raise HTTPException(status_code=404, detail="テンプレートファイルが見つかりません")
    # sheet_nameはExcel内の参照のみに使用されるが、念のためバリデーション済み
    return _read_excel_layout(str(TEMPLATE_PATH), sheet_name)


@router.get("/result-layout/{session_id}")
async def get_result_layout(
    session_id: str = FastAPIPath(..., pattern=r"^[a-f0-9\-]{8,36}$"),
    frame_name: str = Query("frameB", pattern=r"^[a-zA-Z0-9_\-]+$"),
    sheet_name: str = Query("MRC1", pattern=r"^[a-zA-Z0-9_\-]+$"),
):
    """転記済み session の成果物 Excel を表示用 layout として返す。

    転記完了後、フロントエンドが `OUTPUT_DIR/result_{frame_name}_{session_id}.xlsx` を表示するために呼ぶ。
    ファイル名に使う `frame_name`、`session_id`、`sheet_name` は `Path(...).name` でサニタイズし、
    パストラバーサルを避ける。現行形式のファイルが無い場合は、旧形式 `result_{sheet_name}_{session_id}.xlsx` も探す。

    Args:
        session_id (str): 成果物ファイルを特定する session ID。path parameter の pattern でUUID相当へ制限する。
        frame_name (str): 成果物ファイル名に使う frame 名。既定は `frameB`。
        sheet_name (str): 読み取る sheet 名。既定は `MRC1`。

    Returns:
        dict: グリッド表示用 layout。テンプレートと異なり、転記処理で追加された表行も含む。

    Raises:
        HTTPException: 現行形式・旧形式どちらの成果物 Excel も存在しない場合は 404。
        HTTPException: 指定 sheet が Excel 内に無い場合は `_read_excel_layout()` から 404 が伝播する。

    Examples:
        FastAPI endpoint と成果物Excelに依存するため doctest では実行しない。

        >>> response = client.get("/api/result-layout/00000000-0000-0000-0000-000000000000")  # doctest: +SKIP
        >>> "cells" in response.json()  # doctest: +SKIP
        True

    Note:
        layout を返すだけで、Excel の再計算や保存は行わない。式は `_evaluate_supported_formula()` で
        表示用に一部だけ評価する。
    """
    # pathlib.Path.name を使用してサニタイズ（パストラバーサル対策）
    safe_frame = Path(frame_name).name
    safe_session = Path(session_id).name
    safe_sheet = Path(sheet_name).name

    result_path = _resolve_result_path(
        frame_name=safe_frame,
        session_id=safe_session,
        sheet_name=safe_sheet,
    )

    if result_path is None:
        raise HTTPException(status_code=404, detail="転記済みファイルが見つかりません")
    return _read_excel_layout(str(result_path), safe_sheet)


def _resolve_result_path(*, frame_name: str, session_id: str, sheet_name: str) -> Path | None:
    """転記結果ファイルの実体パスを、複数の命名規則から解決する。

    Args:
        frame_name: 様式名。`result_{frame_name}_{session_id}.xlsx` の探索に使う。
        session_id: セッション識別子。通常は Firestore の session_id。
        sheet_name: シート名。旧形式やフォールバック探索に使う。

    Returns:
        Path | None: 見つかった Excel ファイルのパス。見つからない場合は None。

    Example:
        Input:
            frame_name = "frameB"
            session_id = "7bc918c2-5300-40d8-9aa1-a70e05d229a1"
            sheet_name = "MRC1"
        Output:
            Path(".../data/artifacts/form_generation/output/result_frameB_7bc918c2-5300-40d8-9aa1-a70e05d229a1.xlsx")
    """
    candidates: list[Path] = [
        OUTPUT_DIR / f"result_{frame_name}_{session_id}.xlsx",
        OUTPUT_DIR / f"result_{sheet_name}_{session_id}.xlsx",
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    glob_patterns = [
        f"result_{frame_name}_*{session_id}*.xlsx",
        f"result_{sheet_name}_*{session_id}*.xlsx",
        f"result_*{session_id}*.xlsx",
    ]
    globbed: list[Path] = []
    for pattern in glob_patterns:
        globbed.extend(p for p in OUTPUT_DIR.glob(pattern) if p.is_file())

    if not globbed:
        return None

    # フォールバックでも同一 session_id を含む成果物に限定し、同名様式の他案件混入を防ぐ。
    return max(globbed, key=lambda path: path.stat().st_mtime)


def _read_excel_layout(file_path: str, sheet_name: str) -> dict:
    """Excelファイルを読み、フロントエンドのグリッド表示用 JSON を作る。

    `get_template_structure()` と `get_result_layout()` の共通処理。値用 workbook は `data_only=True`、
    式確認用 workbook は `data_only=False` で開く。各セルについて、式があれば formula 文字列を残しつつ、
    対応済みの式だけ `_evaluate_supported_formula()` で表示値へ変換する。

    Args:
        file_path (str): 読み取る Excel ファイル path。
        sheet_name (str): layout 化する sheet 名。

    Returns:
        dict: グリッド表示用 layout。主な key は以下。
        - `sheet_name`: 読み取った sheet 名。
        - `max_row`, `max_col`: sheet の使用範囲。
        - `cells`: 各セルの `row`, `col`, `address`, `value`, `formula`。
        - `merged_cells`: 結合セル範囲。
        - `col_widths`, `row_heights`: 画面表示用に近似変換した幅・高さ。

    Raises:
        HTTPException: 指定 sheet が workbook に無い場合は 404。
        Exception: openpyxl の読み取り例外は捕捉せず伝播する。

    Examples:
        Excel ファイル作成に依存するため doctest では実行しない。

        >>> layout = _read_excel_layout("result.xlsx", "MRC1")  # doctest: +SKIP
        >>> layout["cells"][0]["address"]  # doctest: +SKIP
        'A1'

    Note:
        workbook は `finally` で必ず close する。Excel 式の完全互換評価は目的ではなく、
        画面表示に必要な代表式だけを補助的に評価する。
    """
    wb_values = load_workbook(file_path, data_only=True)
    wb_formulas = load_workbook(file_path, data_only=False)
    try:
        if sheet_name not in wb_values.sheetnames:
            raise HTTPException(status_code=404, detail=f"シート '{sheet_name}' が見つかりません")

        ws = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]
        formula_cache: dict[tuple[str, str], object] = {}

        # セル一覧
        cells = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                formula_value = ws_formulas[cell.coordinate].value
                formula = formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None
                value = cell.value
                if formula:
                    evaluated_value = _evaluate_supported_formula(
                        formula=formula,
                        wb_values=wb_values,
                        wb_formulas=wb_formulas,
                        current_sheet=sheet_name,
                        cache=formula_cache,
                    )
                    if evaluated_value is not None:
                        value = evaluated_value

                cells.append({
                    "row": cell.row,
                    "col": cell.column,
                    "address": cell.coordinate,
                    "value": str(value) if value is not None else None,
                    "formula": formula,
                })

        # 結合セル範囲
        merged = []
        for r in ws.merged_cells.ranges:
            merged.append({
                "start_row": r.min_row,
                "start_col": r.min_col,
                "end_row": r.max_row,
                "end_col": r.max_col,
            })

        # 列幅 (Excel単位 → ピクセル近似)
        col_widths = {}
        for letter, dim in ws.column_dimensions.items():
            idx = column_index_from_string(letter)
            col_widths[str(idx)] = max(int((dim.width or 8) * 8), 40)

        # 行高さ
        row_heights = {}
        for num, dim in ws.row_dimensions.items():
            row_heights[str(num)] = max(int((dim.height or 15) * 1.33), 18)

        return {
            "sheet_name": sheet_name,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "cells": cells,
            "merged_cells": merged,
            "col_widths": col_widths,
            "row_heights": row_heights,
        }
    finally:
        wb_values.close()
        wb_formulas.close()

def _evaluate_supported_formula(
    formula: str,
    wb_values,
    wb_formulas,
    current_sheet: str,
    cache: dict[tuple[str, str], object] | None = None,
) -> float | int | str | None:
    """API表示用に、対応済みのExcel式だけを安全に評価する。

    `_read_excel_layout()` が式セルを見つけたときに呼ぶ入口。先頭の `=` を外して
    `_evaluate_expression()` に渡す。未対応式や評価不能な式は `FormulaEvaluationError` として扱い、
    API表示では例外にせず `None` を返す。

    Args:
        formula (str): Excel セルに入っている式文字列。例: `=SUM(H30:H31)`。
        wb_values: `data_only=True` で開いた workbook。通常セル値の取得に使う。
        wb_formulas: `data_only=False` で開いた workbook。参照先セルが式かどうか確認するために使う。
        current_sheet (str): 式セルがある sheet 名。sheet 省略参照の既定値になる。
        cache (:obj:`dict[tuple[str, str], object]`, optional): セル参照の評価結果 cache。

    Returns:
        float | int | str | None: 評価できた表示値。未対応式や評価不能な式は None。

    Examples:
        workbook に依存するため doctest では実行しない。

        >>> _evaluate_supported_formula("=SUM(H30:H31)", wb_values, wb_formulas, "MRC1")  # doctest: +SKIP
        3

    Note:
        未対応式を 500 エラーにしないため、この関数で `FormulaEvaluationError` を握って None にする。
        それ以外の予期しない例外は `_evaluate_expression()` 側から伝播する可能性がある。
    """
    try:
        return _evaluate_expression(
            expression=formula.strip().lstrip("=").strip(),
            wb_values=wb_values,
            wb_formulas=wb_formulas,
            current_sheet=current_sheet,
            cache={} if cache is None else cache,
        )
    except FormulaEvaluationError:
        return None


def _evaluate_expression(expression: str, wb_values, wb_formulas, current_sheet: str, cache: dict[tuple[str, str], object]):
    """Excel式の一部構文を再帰的に評価する。

    MRC1/MRC2 の画面表示で必要な代表式だけを扱う簡易 evaluator. 数値、文字列、`SUM`、`ROUND`、
    `IF`、`IFERROR`、`&` 連結、四則演算、単一セル参照、範囲参照を評価する。
    それ以外の関数や条件式は `FormulaEvaluationError` にする。

    Args:
        expression (str): `=` を外した Excel 式。例: `SUM(H30:H31)`。
        wb_values: `data_only=True` workbook。
        wb_formulas: `data_only=False` workbook。
        current_sheet (str): sheet 省略参照の既定 sheet。
        cache (dict[tuple[str, str], object]): セル参照の評価結果 cache。循環参照検出にも使う。

    Returns:
        float | int | str: 評価結果。空 expression は空文字、整数相当の float は int に寄せる。

    Raises:
        FormulaEvaluationError: 未対応関数、未対応条件、割り算の0除算、数値変換不能、未対応 expression の場合。

    Examples:
        workbook 参照を含む式は doctest では実行しない。

        >>> _evaluate_expression('"A"', None, None, "MRC1", {})
        'A'
        >>> _evaluate_expression('1+2', None, None, "MRC1", {})
        3

    Note:
        この evaluator はExcel互換を目指したものではない。画面表示に必要な式だけを最小限処理する。
    """
    expression = _strip_wrapping_parentheses(expression.strip())
    if not expression:
        return ""

    if expression.startswith('"') and expression.endswith('"'):
        return expression[1:-1]

    if _NUMBER_RE.match(expression):
        number = float(expression)
        return int(number) if number.is_integer() else number

    function_match = _match_function(expression)
    if function_match:
        func_name, raw_args = function_match
        args = _split_formula_args(raw_args)
        upper_name = func_name.upper()
        if upper_name == "SUM":
            total = 0.0
            for arg in args:
                total += _coerce_number(_evaluate_sum_arg(arg, wb_values, wb_formulas, current_sheet, cache))
            return int(total) if total.is_integer() else total
        if upper_name == "ROUND" and len(args) == 2:
            value = _coerce_number(
                _evaluate_expression(args[0], wb_values, wb_formulas, current_sheet, cache),
                strict=True,
            )
            digits = int(
                _coerce_number(
                    _evaluate_expression(args[1], wb_values, wb_formulas, current_sheet, cache),
                    strict=True,
                )
            )
            rounded = round(value, digits)
            return int(rounded) if float(rounded).is_integer() else rounded
        if upper_name == "IF" and len(args) == 3:
            return _evaluate_expression(args[1], wb_values, wb_formulas, current_sheet, cache) if _evaluate_condition(args[0], wb_values, wb_formulas, current_sheet, cache) else _evaluate_expression(args[2], wb_values, wb_formulas, current_sheet, cache)
        if upper_name == "IFERROR" and len(args) == 2:
            try:
                return _evaluate_expression(args[0], wb_values, wb_formulas, current_sheet, cache)
            except FormulaEvaluationError:
                return _evaluate_expression(args[1], wb_values, wb_formulas, current_sheet, cache)
        raise FormulaEvaluationError(f"unsupported function: {func_name}")

    concat_parts = _split_top_level(expression, "&")
    if len(concat_parts) > 1:
        return "".join(
            _stringify_formula_value(_evaluate_expression(part, wb_values, wb_formulas, current_sheet, cache))
            for part in concat_parts
        )

    for operator in ("+", "-", "*", "/"):
        split_result = _split_binary_expression(expression, operator)
        if split_result is None:
            continue
        left_expr, right_expr = split_result
        left = _coerce_number(
            _evaluate_expression(left_expr, wb_values, wb_formulas, current_sheet, cache),
            strict=True,
            blank_as_zero=True,
        )
        right = _coerce_number(
            _evaluate_expression(right_expr, wb_values, wb_formulas, current_sheet, cache),
            strict=True,
            blank_as_zero=True,
        )
        if operator == "+":
            result = left + right
        elif operator == "-":
            result = left - right
        elif operator == "*":
            result = left * right
        else:
            if right == 0:
                raise FormulaEvaluationError("division by zero")
            result = left / right
        return int(result) if float(result).is_integer() else result

    cell_match = _CELL_REF_RE.match(expression)
    if cell_match:
        ref_sheet = _normalize_sheet_name(cell_match.group("sheet")) or current_sheet
        ref_cell = _normalize_cell_address(cell_match.group("cell"))
        return _get_cell_value(ref_sheet, ref_cell, wb_values, wb_formulas, cache)

    raise FormulaEvaluationError(f"unsupported expression: {expression}")


def _get_worksheet(workbook, sheet_name: str):
    """式評価用に worksheet を取得する。存在しない sheet は評価不能として扱う。

    openpyxl の `workbook[sheet_name]` は sheet が無いと `KeyError` を投げる。式評価中の
    `KeyError` は `_evaluate_supported_formula()` で捕捉されず API 全体が 500 になるため、
    この helper で `FormulaEvaluationError` へ変換し、「失敗した式だけ表示値 None」の
    設計に合わせる。予期しない `KeyError` まで握り潰さないよう、入口で広く捕捉する方式は取らない。

    Args:
        workbook: `wb_values` または `wb_formulas`。
        sheet_name (str): 参照先 sheet 名。

    Returns:
        Worksheet: 取得した worksheet。

    Raises:
        FormulaEvaluationError: sheet が workbook に存在しない場合。

    Examples:
        workbook に依存するため doctest では実行しない。

        >>> _get_worksheet(wb_values, "MRC1")  # doctest: +SKIP
        <Worksheet "MRC1">

    Note:
        `_read_excel_layout()` の表示対象 sheet の存在確認（404 を返す仕様）とは役割が異なる。
        こちらは式の参照先だけを対象にする。
    """
    if sheet_name not in workbook.sheetnames:
        raise FormulaEvaluationError(f"missing sheet: {sheet_name}")
    return workbook[sheet_name]


def _evaluate_sum_arg(arg: str, wb_values, wb_formulas, current_sheet: str, cache: dict[tuple[str, str], object]):
    """SUM の引数1つを数値として評価する。

    `SUM(G39:G46)` のような範囲参照なら範囲内の各セルを合計し、単一 expression なら
    `_evaluate_expression()` の結果を数値へ寄せる。

    Args:
        arg (str): SUM の引数1つ。範囲参照または通常 expression。
        wb_values: `data_only=True` workbook。
        wb_formulas: `data_only=False` workbook。
        current_sheet (str): sheet 省略参照の既定 sheet。
        cache (dict[tuple[str, str], object]): セル参照評価 cache。

    Returns:
        float: SUM 引数として使う数値。

    Raises:
        FormulaEvaluationError: 参照先の式評価や数値変換で失敗した場合。

    Examples:
        workbook 範囲参照に依存するため doctest では実行しない。

        >>> _evaluate_sum_arg("G39:G46", wb_values, wb_formulas, "MRC2", {})  # doctest: +SKIP
        36.0

    Note:
        空セルや非数値は `_coerce_number()` の既定挙動により 0.0 として扱う。
    """
    range_match = _RANGE_REF_RE.match(arg.strip())
    if not range_match:
        return _evaluate_expression(arg, wb_values, wb_formulas, current_sheet, cache)

    target_sheet = _normalize_sheet_name(range_match.group("sheet")) or current_sheet
    ws = _get_worksheet(wb_values, target_sheet)
    min_col, min_row, max_col, max_row = range_boundaries(
        f"{_normalize_cell_address(range_match.group('start'))}:{_normalize_cell_address(range_match.group('end'))}"
    )
    total = 0.0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            total += _coerce_number(
                _get_cell_value(target_sheet, cell.coordinate, wb_values, wb_formulas, cache)
            )
    return total


def _evaluate_condition(expression: str, wb_values, wb_formulas, current_sheet: str, cache: dict[tuple[str, str], object]) -> bool:
    """IF 条件部の等価比較だけを評価する。

    現行テンプレートで必要な `A=B` 形式の条件だけを扱う。左右を `_evaluate_expression()` で評価し、
    `_stringify_formula_value()` で文字列化して比較する。

    Args:
        expression (str): IF の条件部。例: `$C$4=""`。
        wb_values: `data_only=True` workbook。
        wb_formulas: `data_only=False` workbook。
        current_sheet (str): sheet 省略参照の既定 sheet。
        cache (dict[tuple[str, str], object]): セル参照評価 cache。

    Returns:
        bool: 左右が文字列として等しければ True。

    Raises:
        FormulaEvaluationError: `=` で top-level に2分割できない条件式の場合。

    Examples:
        >>> _evaluate_condition('"A"="A"', None, None, "MRC1", {})
        True

    Note:
        `>`、`<`、`>=`、`<=` などの比較演算子は現行実装では未対応。
    """
    parts = _split_top_level(expression, "=")
    if len(parts) != 2:
        raise FormulaEvaluationError(f"unsupported condition: {expression}")
    left = _evaluate_expression(parts[0], wb_values, wb_formulas, current_sheet, cache)
    right = _evaluate_expression(parts[1], wb_values, wb_formulas, current_sheet, cache)
    return _stringify_formula_value(left) == _stringify_formula_value(right)


def _get_cell_value(sheet_name: str, cell_address: str, wb_values, wb_formulas, cache: dict[tuple[str, str], object]):
    """セル参照を解決し、必要なら参照先の式も評価する。

    `_evaluate_expression()` が単一セル参照を見つけたときに呼ぶ helper。
    `wb_values` から表示値を読み、`wb_formulas` で同じセルに式が入っているかを確認する。
    参照先も式なら再帰的に評価し、cache に保存する。評価中マーカーを使って循環参照を検出する。

    Args:
        sheet_name (str): 参照先 sheet 名。
        cell_address (str): 参照先セル番地。`$` 付きでもよい。
        wb_values: `data_only=True` workbook。
        wb_formulas: `data_only=False` workbook。
        cache (dict[tuple[str, str], object]): 評価済みセル cache と循環参照検出用 state。

    Returns:
        object: セルの値、または式評価結果。

    Raises:
        FormulaEvaluationError: 循環参照を検出した場合、または参照先式の評価に失敗した場合。

    Examples:
        workbook セル参照に依存するため doctest では実行しない。

        >>> _get_cell_value("MRC1", "C4", wb_values, wb_formulas, {})  # doctest: +SKIP
        2025

    Note:
        式評価中に失敗したセルは cache から取り除き、次回評価が壊れた中間状態を再利用しないようにする。
    """
    key = (sheet_name, _normalize_cell_address(cell_address))
    if key in cache:
        cached = cache[key]
        if cached is _IN_PROGRESS:
            raise FormulaEvaluationError(f"circular reference: {sheet_name}!{cell_address}")
        return cached

    ws_values = _get_worksheet(wb_values, sheet_name)
    ws_formulas = _get_worksheet(wb_formulas, sheet_name)
    value = ws_values[cell_address].value
    formula = ws_formulas[cell_address].value
    if not (isinstance(formula, str) and formula.startswith("=")):
        return value

    cache[key] = _IN_PROGRESS
    try:
        evaluated = _evaluate_expression(formula.lstrip("="), wb_values, wb_formulas, sheet_name, cache)
        if evaluated is None and value is not None:
            evaluated = value
        cache[key] = evaluated
        return evaluated
    except Exception:
        cache.pop(key, None)
        raise


def _match_function(expression: str) -> tuple[str, str] | None:
    """expression が `NAME(args)` 形式の関数呼び出しか判定する。

    `_evaluate_expression()` が `SUM(...)` や `IF(...)` などを処理する前段で使う。
    正規表現で関数名と引数文字列を取り出し、さらに `_split_formula_args()` で括弧や文字列の対応が
    壊れていないことを確認する。

    Args:
        expression (str): 判定対象の式文字列。

    Returns:
        tuple[str, str] | None: 関数形式なら `(関数名, raw_args)`。関数形式でない、または引数構文が壊れている場合は None。

    Examples:
        >>> _match_function("SUM(A1:A3)")
        ('SUM', 'A1:A3')
        >>> _match_function("A1+1") is None
        True

    Note:
        この関数は関数名が対応済みかどうかまでは見ない。対応可否は `_evaluate_expression()` 側で判定する。
    """
    match = _FUNCTION_RE.match(expression)
    if not match:
        return None
    name = match.group("name")
    args = match.group("args")
    if _split_formula_args(args) is None:
        return None
    return name, args


def _split_formula_args(raw_args: str) -> list[str] | None:
    """Excel関数の引数文字列を top-level のカンマで分割する。

    `IF(A1="", "N", A1)` のように、引数内に括弧や文字列がある場合でも、括弧の外・文字列の外にある
    カンマだけを分割点にする。括弧やダブルクォートが閉じていない場合は None を返す。

    Args:
        raw_args (str): 関数呼び出しの括弧内文字列。

    Returns:
        list[str] | None: 分割済み引数配列。構文が壊れている場合は None。

    Examples:
        >>> _split_formula_args('A1,"B,C",SUM(D1:D2)')
        ['A1', '"B,C"', 'SUM(D1:D2)']
        >>> _split_formula_args('A1,(B1') is None
        True

    Note:
        Excel の区切り文字は現行テンプレート前提でカンマのみ扱う。セミコロン区切り locale は未対応。
    """
    args: list[str] = []
    current = []
    depth = 0
    in_string = False
    for char in raw_args:
        if char == '"':
            in_string = not in_string
        elif not in_string and char == "(":
            depth += 1
        elif not in_string and char == ")":
            depth -= 1
            if depth < 0:
                return None
        if char == "," and not in_string and depth == 0:
            args.append("".join(current).strip())
            current = []
            continue
        current.append(char)
    if depth != 0 or in_string:
        return None
    args.append("".join(current).strip())
    return args


def _split_top_level(expression: str, delimiter: str) -> list[str]:
    """括弧や文字列の内側を避けて、top-level の delimiter で式を分割する。

    文字列連結 `&`、条件式 `=` などを分割するときに使う helper。たとえば `"A&B"&C1` は
    文字列内の `&` では分割せず、外側の `&` だけで分ける。

    Args:
        expression (str): 分割対象の式。
        delimiter (str): 分割に使う1文字の区切り記号。

    Returns:
        list[str]: 分割結果。delimiter が無ければ元 expression だけの1要素配列。

    Examples:
        >>> _split_top_level('"A&B"&C1', '&')
        ['"A&B"', 'C1']
        >>> _split_top_level('IF(A1="",B1,C1)', ',')
        ['IF(A1="",B1,C1)']

    Note:
        この関数は括弧の不整合をエラーにしない。分割位置を決めるための軽量 parser として使う。
    """
    parts: list[str] = []
    current = []
    depth = 0
    in_string = False
    for char in expression:
        if char == '"':
            in_string = not in_string
        elif not in_string and char == "(":
            depth += 1
        elif not in_string and char == ")":
            depth -= 1
        if char == delimiter and not in_string and depth == 0:
            parts.append("".join(current).strip())
            current = []
            continue
        current.append(char)
    parts.append("".join(current).strip())
    return parts


def _split_binary_expression(expression: str, operator: str) -> tuple[str, str] | None:
    """四則演算を、最上位の演算子位置で左辺・右辺に分ける。

    `_evaluate_expression()` が `A1+B1` や `E39*F39` を評価するときに使う。
    右から左へ走査し、括弧や文字列の中にある演算子を無視する。先頭の `+`/`-` は符号として扱い、
    二項演算子とは見なさない。

    Args:
        expression (str): 分割対象の式。
        operator (str): 探す演算子。`+`、`-`、`*`、`/` のいずれか。

    Returns:
        tuple[str, str] | None: 分割できた場合は `(left_expr, right_expr)`。対象演算子が無ければ None。

    Examples:
        >>> _split_binary_expression("A1+B1", "+")
        ('A1', 'B1')
        >>> _split_binary_expression("-1", "-") is None
        True

    Note:
        Excel の演算子優先順位は `_evaluate_expression()` が `+`、`-`、`*`、`/` の順にこの関数を呼ぶことで
        簡易的に表現している。
    """
    depth = 0
    in_string = False
    for index in range(len(expression) - 1, -1, -1):
        char = expression[index]
        if char == '"':
            in_string = not in_string
            continue
        if in_string:
            continue
        if char == ")":
            depth += 1
            continue
        if char == "(":
            depth -= 1
            continue
        if depth != 0 or char != operator:
            continue
        if operator in "+-" and index == 0:
            continue
        left = expression[:index].strip()
        right = expression[index + 1:].strip()
        if not left or not right:
            continue
        return left, right
    return None


def _strip_wrapping_parentheses(expression: str) -> str:
    """式全体を包むだけの外側括弧を取り除く。

    `((A1+B1))` のように式全体を囲む括弧は評価上不要なので外す。一方、`(A1+B1)*C1` のように
    外側括弧の後に別演算が続く場合は、式全体を包んでいないため外さない。

    Args:
        expression (str): 括弧を取り除く対象の式。

    Returns:
        str: 外側括弧を取り除いた式。

    Examples:
        >>> _strip_wrapping_parentheses("((A1+B1))")
        'A1+B1'
        >>> _strip_wrapping_parentheses("(A1+B1)*C1")
        '(A1+B1)*C1'

    Note:
        括弧が不均衡な場合は無理に外さず、後続の評価処理に任せる。
    """
    while expression.startswith("(") and expression.endswith(")"):
        depth = 0
        balanced = True
        for index, char in enumerate(expression):
            if char == "(":
                depth += 1
            elif char == ")":
                depth -= 1
                if depth == 0 and index != len(expression) - 1:
                    balanced = False
                    break
        if not balanced or depth != 0:
            break
        expression = expression[1:-1].strip()
    return expression


def _normalize_sheet_name(raw_sheet_name: str | None) -> str | None:
    """Excel式内の sheet 名表記から、外側の引用符を取り除く。

    クロスシート参照では `'MRC1'!C4` のように sheet 名がシングルクォートで囲まれることがある。
    この helper は参照処理で使いやすいように、空値は None のまま、文字列は trim して外側の `'` を外す。

    Args:
        raw_sheet_name (:obj:`str`, optional): 正規化前の sheet 名。例: `'MRC1'`。

    Returns:
        str | None: 正規化後の sheet 名。入力が空なら None。

    Examples:
        >>> _normalize_sheet_name("'MRC1'")
        'MRC1'
        >>> _normalize_sheet_name(None) is None
        True

    Note:
        sheet 名の存在確認はここでは行わない。実際の workbook 参照時に行われる。
    """
    if not raw_sheet_name:
        return None
    return raw_sheet_name.strip().strip("'")


def _normalize_cell_address(address: str) -> str:
    """Excelセル参照から絶対参照記号 `$` を外し、大文字にそろえる。

    openpyxl のセルアクセスでは `C4` の形にそろえると扱いやすい。Excel式では `$C$4` や `c4` のような
    表記があり得るため、この helper で正規化する。

    Args:
        address (str): Excel セル番地。例: `$C$4`、`c4`。

    Returns:
        str: 正規化済みセル番地。例: `C4`。

    Examples:
        >>> _normalize_cell_address("$c$4")
        'C4'

    Note:
        範囲参照全体ではなく、単一セル番地だけを対象にする。
    """
    return address.replace("$", "").upper()


def _stringify_formula_value(value) -> str:
    """式評価結果を Excel の文字列比較・連結で使いやすい文字列にする。

    IF 条件の左右比較や `&` 連結では、評価済み値を文字列化する必要がある。
    None は空文字にし、`2025.0` のような整数相当 float は `2025` に寄せる。

    Args:
        value: 文字列化する式評価結果。

    Returns:
        str: Excel表示に近づけた文字列。

    Examples:
        >>> _stringify_formula_value(None)
        ''
        >>> _stringify_formula_value(2025.0)
        '2025'

    Note:
        日付や通貨などのExcel書式までは再現しない。式表示に必要な最小限の文字列化だけを行う。
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _coerce_number(value, strict: bool = False, blank_as_zero: bool = True) -> float:
    """式評価結果を数値演算用の float に変換する。

    `SUM` や四則演算では、セル値や式評価結果を数値として扱う必要がある。
    空値は既定で 0.0 とし、文字列はカンマを外して float 変換する。`strict=True` の場合、
    変換できない値は `FormulaEvaluationError` にする。

    Args:
        value: 数値化したい値。int, float, 数値文字列, 空値などを想定する。
        strict (bool): True の場合、非数値を 0.0 にせず例外にする。
        blank_as_zero (bool): 空値を 0.0 として扱うか。`strict=True` かつ False の場合は空値も例外。

    Returns:
        float: 数値化した値。非strictで変換できない場合は 0.0。

    Raises:
        FormulaEvaluationError: `strict=True` で空値または非数値を受け取った場合。

    Examples:
        >>> _coerce_number("1,234")
        1234.0
        >>> _coerce_number("abc")
        0.0

    Note:
        Excel のエラー値やパーセント書式などは扱わない。現行テンプレートの表示式に必要な数値化だけを行う。
    """
    if value in (None, ""):
        if strict and not blank_as_zero:
            raise FormulaEvaluationError("empty value")
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        if strict:
            raise FormulaEvaluationError(f"non numeric value: {value}")
        return 0.0


_IN_PROGRESS = object()